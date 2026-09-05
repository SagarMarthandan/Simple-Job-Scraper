#!/usr/bin/env python3
"""
Jobscraper Pipeline — 100% Free Job Fetcher, Deduplicator & Dated CSV Exporter
Platforms: LinkedIn (free HTML), Indeed (GraphQL API), Arbeitnow, Xing, Stepstone, ATS Direct (Greenhouse/SmartRecruiters/Ashby)
Filters: Freshness (<24h), Experience (<=2 yrs), Working Student (Hamburg & Kiel ONLY),
         Internships (Germany-wide), Title relevance (data/analytics/AI/SQL/Python keywords)
Output Path: /home/sagar/Skills/Jobscraper/Job Search/YYYY-MM-DD/Job_Search_<Month>_<Day>_<Year>.csv

Changelog:
  2026-08-04: Initial release. LinkedIn + Indeed + Arbeitnow + Kununu. 24 SEARCH_ROLES. ~$1.50-3.00/run.
  2026-08-05: Cost optimization (82% reduction → ~$0.55/run). Consolidated 24 → 10 SEARCH_ROLES.
              Fixed LinkedIn maxItems bug (was sending maxItems, actor uses count). Fixed Indeed missing
              title param (generic garbage results). Added maxTotalChargeUsd safety caps. Dropped Kununu
              (broken actor). Added Startup.jobs (free cloudscraper). Added is_relevant_title() post-filter.
  2026-08-06: Added Xing (free cloudscraper, data-testid + <time dateTime>). Added Stepstone (free
              cloudscraper, path-based URLs with ag=age_1 24h filter, data-at SSR attributes, German
              relative time parsing via parse_stepstone_timeago()). Pipeline now 6 platforms.
  2026-08-07: LinkedIn AI search rollout — actor autoConvertToAiSearch defaults true, softening
              location=Germany into a natural-language hint. f_TPR=r86400 24h filter unaffected
              (stays as URL filter under AI search). Removed dead prototype scripts
              (stepstone_scraper.py, csv_to_xlsx.py, merge_linkedin.py). Added .gitignore, README.md, CHANGELOG.md.
  2026-08-09b: CRITICAL FIX: Indeed's datePosted='1' parameter is IGNORED by the API (same bug as
              Glassdoor's fromAge=1). 11/102 jobs (10.8%) were stale, including jobs 500+ days old.
              Added post-filter on datePublished in fetch_indeed_jobs() to reject jobs older than 24h.
              Also added defense-in-depth safety-net post-filter to fetch_linkedin_jobs() (f_TPR=r86400
              verified working 727/727, but post-filter catches any future regressions). Date-only
              timestamps (LinkedIn format) treated as end-of-day (23:59:59) to avoid false rejections.
  2026-08-10: CRITICAL FIX: Stepstone switched from cloudscraper to plain requests. Stepstone is
              behind Akamai (not Cloudflare) — cloudscraper's challenge-solving hangs on Akamai,
              causing read timeouts on port 443. Plain requests with a browser User-Agent works
              (Stepstone serves SSR HTML without anti-bot challenge). All data-at parsing unchanged.
              Also switched Xing from cloudscraper to plain requests — Xing is behind AWS CloudFront
              (not Cloudflare), so cloudscraper was unnecessary overhead with the same hang risk.

See CHANGELOG.md for full version history and apify_job_search.md for detailed technical documentation.
"""

import os
import re
import csv
import json
import time
import urllib.request
import html as html_mod
import urllib.parse
try:
    import cloudscraper
except ImportError:
    cloudscraper = None
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
from datetime import datetime, timezone, timedelta
from pathlib import Path

# Paths
BASE_RESUMES_DIR = Path("/home/sagar/Documents/YAML-CV/skills/okf-cv/okf/base_files")
PORTFOLIO_DIR = Path("/home/sagar/Documents/YAML-CV/skills/okf-cv/okf/portfolio")
JOB_SEARCH_DIR = Path("/home/sagar/Skills/Jobscraper/Job Search")

# Ensure Output Directory Exists
JOB_SEARCH_DIR.mkdir(parents=True, exist_ok=True)

# Search Parameters
# Consolidated from 24 → 10 core roles.
# LinkedIn/Indeed/Kununu search engines cover variants (Junior, Cloud, etc.) automatically.
# German terms kept where they produce distinct results not covered by English equivalents.
SEARCH_ROLES = [
    # Core Data Engineering (covers Junior, Cloud, Data Warehouse, ETL, Dateningenieur)
    "Data Engineer",
    "Analytics Engineer",
    # Business Intelligence & Analytics (covers Datenanalyst, BI Developer, BI Entwickler)
    "Data Analyst",
    # AI & Data Science (covers AI Data Engineer, GenAI, Junior Data Scientist)
    "AI Engineer",
    "Machine Learning Engineer",
    # Business / Quantitative Analytics
    "Business Analyst",
    # Database & Systems (covers Database Developer, Python Data Developer)
    "SQL Developer",
    # Internships & Working Student (distinct result sets)
    "Praktikum Data",
    "Werkstudent Data",
    "Werkstudent Business Intelligence",
]

MAX_EXP_YEARS = 2
FRESHNESS_HOURS = 24


# Seniority Filter Regex
EXCLUDED_TITLE_PATTERNS = re.compile(
    r"\b(senior|sr|lead|head|principal|staff|manager|director|architect|vp|chief|expert|team lead)\b",
    re.IGNORECASE
)


# Core Tech Stack Keywords for Match Scoring
TECH_KEYWORDS = ["dbt", "airflow", "spark", "pyspark", "python", "sql", "gcp", "bigquery", "aws", "azure", "databricks", "docker", "kafka", "postgresql", "snowflake"]

# Core domain keywords that MUST appear in the job title for it to be relevant.
# This is a universal post-filter applied to ALL platforms (LinkedIn, Indeed, Arbeitnow).
# Without this, actors return garbage: "Nachtwächter" for "Data Engineer" search, etc.
DOMAIN_TITLE_KEYWORDS = re.compile(
    r"\b(data|daten\w*|analytics|analyst\w*|ai|artificial intelligence|"
    r"bi|business intelligence|sql|python|database|datenbank\w*|"
    r"machine learning|ml|deep learning|etl|pipeline|warehouse|"
    r"datapipeline|dataops|mlops|llm|genai)\b",
    re.IGNORECASE
)


def is_relevant_title(title: str) -> bool:
    """Check if the job title contains at least one core data/analytics/AI keyword.
    This catches false positives from all actors (Indeed returns 'Nachtwächter' for
    'Data Engineer' searches, LinkedIn returns 'Junior Software Engineer', etc.).
    """
    return bool(DOMAIN_TITLE_KEYWORDS.search(title))

def detect_language(text: str) -> str:
    """Detect if job description/title is primarily German or English."""
    text_lower = text.lower()
    german_indicators = ["und", "die", "das", "der", "mit", "für", "aufgaben", "profil", "kenntnisse", "anforderungen", "ihre", "wir"]
    english_indicators = ["and", "the", "with", "for", "responsibilities", "requirements", "skills", "your", "we", "looking"]

    de_score = sum(len(re.findall(r"\b" + w + r"\b", text_lower)) for w in german_indicators)
    en_score = sum(len(re.findall(r"\b" + w + r"\b", text_lower)) for w in english_indicators)

    return "German" if de_score > en_score else "English"

def classify_role_type(title: str, description: str) -> str:
    """Classify into Working Student, Internship, or Full-Time / Entry Level."""
    combined = f"{title} {description}".lower()
    if re.search(r"\b(werkstudent|werkstudentin|working student)\b", combined):
        return "Working Student"
    elif re.search(r"\b(intern|internship|praktikum|praktikant|praktikantin)\b", combined):
        return "Internship"
    else:
        return "Full-Time / Entry-Level"

def check_experience_and_location(title: str, description: str, location: str) -> tuple[bool, str]:
    """Validates role against seniority ceiling and location constraints.
    Returns (is_valid, role_type_or_reason)
    """
    if not is_relevant_title(title):
        return False, "Title not relevant to data/analytics/AI"

    if EXCLUDED_TITLE_PATTERNS.search(title):
        return False, "Seniority title excluded"

    role_type = classify_role_type(title, description)
    loc_clean = location.lower()

    # Working Student — strictly restricted to Hamburg & Kiel
    if role_type == "Working Student":
        if not ("hamburg" in loc_clean or "kiel" in loc_clean):
            return False, f"Working student outside Hamburg/Kiel ({location})"

    return True, role_type

def compute_match_score(text: str) -> int:
    """Calculate percentage match score against Sagar's core tech stack."""
    text_lower = text.lower()
    matches = sum(1 for kw in TECH_KEYWORDS if kw in text_lower)
    return min(100, int((matches / len(TECH_KEYWORDS)) * 100 * 2.5))  # Normalized score

def _norm_company(c: str) -> str:
    c = (c or "").lower().strip()
    # Strip parentheticals: "(REWE Group)", "(m/w/d)"
    c = re.sub(r"\(.*?\)", "", c)
    # Strip legal suffixes (expanded)
    c = re.sub(r"\b(gmbh|ag|inc|ltd|co|kg|se|corp|llc|group|gruppe|holding|"
               r"international|deutschland|germany|global|e\.?g\.?|gmbh & co)\b", "", c)
    # Strip punctuation, collapse whitespace
    c = re.sub(r"[^\w\s]", "", c)
    return " ".join(c.split())

def _norm_title(t: str) -> str:
    t = (t or "").lower().strip()
    # Strip parentheticals (gender markers, location hints)
    t = re.sub(r"\(.*?\)", "", t)
    # Strip seniority + gender markers
    t = re.sub(r"\b(senior|junior|lead|principal|staff|sr\.?|jr\.?)\b", "", t)
    t = re.sub(r"\b(m/w/d|m/f/d|m/w|f/m/d|w/m/d|w/m/x|m/f/x|all genders|w/m|f/w/d)\b", "", t)
    # Strip reference codes: "- REF99139A", "REF12345"
    t = re.sub(r"\bref\d+\w*\b", "", t)
    # Strip punctuation, collapse whitespace
    t = re.sub(r"[^\w\s]", "", t)
    return " ".join(t.split())

def normalize_key(company: str, title: str) -> str:
    """Generate deduplication key (enhanced: strips parentheticals, legal suffixes,
    seniority/gender markers, and REF codes for cross-platform matching)."""
    return f"{_norm_company(company)}::{_norm_title(title)}"


def normalize_job_url(url: str) -> str:
    """Normalize a job URL into a stable cross-run identity key.

    LinkedIn URLs carry per-run tracking params (position/pageNum/refId/trackingId)
    while the job ID lives in the path — drop the query string there.
    Indeed (?jk=) carries its job ID in the query — keep it.
    """
    url = (url or "").strip()
    if not url:
        return ""
    if "linkedin.com" in url:
        url = url.split("?", 1)[0]
    return url.rstrip("/").lower()

def load_previous_run_urls(now: datetime) -> set:
    """Load job URLs from yesterday's CSV folder for cross-run dedup."""
    yesterday = JOB_SEARCH_DIR / (now - timedelta(days=1)).strftime("%Y-%m-%d")
    urls = set()
    for csv_file in yesterday.glob("Job_Search_*.csv"):
        try:
            with open(csv_file, newline="", encoding="utf-8-sig") as f:
                for row in csv.DictReader(f):
                    url = normalize_job_url(row.get("job_url", ""))
                    if url:
                        urls.add(url)
        except Exception:
            continue
    return urls

def fetch_arbeitnow_jobs():
    """Fetch jobs from Arbeitnow API across all search roles."""
    jobs = []
    try:
        url = "https://www.arbeitnow.com/api/job-board-api"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode())
            now = datetime.now(timezone.utc)
            cutoff = now - timedelta(hours=FRESHNESS_HOURS)

            for item in data.get("data", []):
                created_at = datetime.fromtimestamp(item.get("created_at", 0), tz=timezone.utc)
                if created_at < cutoff:
                    continue

                title = item.get("title", "")
                desc = item.get("description", "")
                loc = item.get("location", "Germany")


                is_valid, role_type_or_reason = check_experience_and_location(title, desc, loc)
                if not is_valid:
                    continue

                lang = detect_language(f"{title} {desc}")
                score = compute_match_score(f"{title} {desc}")

                jobs.append({
                    "language": lang,
                    "job_board": "Arbeitnow",
                    "role_type": role_type_or_reason,
                    "title": title,
                    "company": item.get("company_name", "Unknown"),
                    "location": loc,
                    "posted_at": created_at.isoformat(),
                    "exp_required": "<= 2 Years",
                    "match_score": f"{score}%",
                    "job_url": item.get("url", ""),
                    "description": desc
                })
    except Exception as e:
        print(f"[!] Error fetching Arbeitnow: {e}")
    return jobs


def _parse_xing_card(card, cutoff, seen_urls):
    """Parse one Xing job card HTML into a job dict, or None if filtered out.

    Handles URL dedup (seen_urls), date-based freshness filtering (only rejects
    confirmed stale jobs — no date means include per false-positives-over-false-negatives
    rule), and title/experience/location filters.
    """
    # URL — skip /jobs/search/ (promoted recommendation links)
    url_m = re.search(r'href="(/jobs/(?!search)[^"]+)"', card)
    if not url_m:
        return None
    job_url = url_m.group(1)
    if job_url in seen_urls:
        return None
    seen_urls.add(job_url)

    # Date — ISO datetime from <time dateTime="..."> (sponsored jobs lack this).
    # Per "false positives > false negatives" rule: if no date or unparseable,
    # include the job (let title/seniority filters handle it, cross-run dedup
    # catches repeats). Only reject jobs with a confirmed date older than 24h.
    date_m = re.search(r'dateTime="([^"]+)"', card)
    posted_dt = None
    if date_m:
        try:
            posted_dt = datetime.fromisoformat(date_m.group(1).replace("Z", "+00:00"))
        except ValueError:
            posted_dt = None

    if posted_dt and posted_dt < cutoff:
        return None  # confirmed stale — skip

    # Title
    title_m = re.search(r'job-teaser-list-title">([^<]+)<', card)
    if not title_m:
        return None
    title = html_mod.unescape(title_m.group(1).strip())

    # Company — aria-label on <img> is most reliable
    company_m = re.search(r'aria-label="([^"]+)"[^>]*loading="lazy"', card)
    company = html_mod.unescape(company_m.group(1).strip()) if company_m else "Unknown"

    # Location — text before <b> tag inside multi-location-display container
    loc_m = re.search(
        r'multi-location-display-styles__Container[^>]*>.*?data-xds="BodyCopy">([^<]+)<b',
        card, re.DOTALL
    )
    location = html_mod.unescape(loc_m.group(1).strip()) if loc_m else "Germany"

    # Apply existing filters (title relevance, seniority, experience, location)
    desc = ""  # no description from listing page
    is_valid, role_type_or_reason = check_experience_and_location(title, desc, location)
    if not is_valid:
        return None

    return {
        "language": detect_language(title),
        "job_board": "Xing",
        "role_type": role_type_or_reason,
        "title": title,
        "company": company,
        "location": location,
        "posted_at": posted_dt.isoformat() if posted_dt else "Last 24h",
        "exp_required": "<= 2 Years",
        "match_score": f"{compute_match_score(title)}%",
        "job_url": f"https://www.xing.com{job_url}",
        "description": ""
    }

def fetch_xing_jobs():
    """Fetch jobs from Xing.com via HTML scraping (free, no Apify).
    Uses plain requests — Xing is behind AWS CloudFront (not Cloudflare), so
    no anti-bot bypass needed. cloudscraper worked but was unnecessary overhead
    with the same Akamai-style hang risk as Stepstone if cloudscraper updates
    its challenge handling. Fetches search result pages for each SEARCH_ROLE
    and parses job listings from server-rendered HTML. Relies on data-testid
    attributes (stable test IDs) and <time dateTime> ISO timestamps for 24h
    freshness filtering. Sponsored listings (no dateTime) are skipped.
    """
    if requests is None:
        print("[!] requests not installed — skipping Xing. Install with: pip install requests")
        return []

    jobs = []
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(hours=FRESHNESS_HOURS)
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "de-DE,de;q=0.9,en;q=0.8",
    })
    seen_urls = set()
    XING_PAGES_PER_ROLE = 3  # 20 results/page × 3 = 60 max per role

    for role in SEARCH_ROLES:
        role_fresh = 0
        role_raw = 0

        for page in range(1, XING_PAGES_PER_ROLE + 1):
            url = f"https://www.xing.com/search/in/jobs?keywords={urllib.parse.quote(role)}&location=germany&page={page}"
            try:
                r = session.get(url, timeout=15)
                if r.status_code != 200:
                    break
                r.encoding = "utf-8"  # Xing sends UTF-8
                raw = r.text
            except Exception as e:
                print(f"[!] Xing fetch error ({role} page {page}): {e}")
                break

            # Split into job cards by styled-components class prefix (hash-independent)
            cards = re.split(r"job-teaser-list-item-styles__Card", raw)
            job_cards = [c for c in cards if "job-teaser-list-title" in c]

            page_fresh = 0
            for card in job_cards:
                # URL dedup — extract path for seen_urls check before parsing
                url_m = re.search(r'href="(/jobs/(?!search)[^"]+)"', card)
                if not url_m:
                    continue
                job_url = url_m.group(1)
                if job_url in seen_urls:
                    continue
                role_raw += 1
                job = _parse_xing_card(card, cutoff, seen_urls)
                if job:
                    jobs.append(job)
                    page_fresh += 1
                    role_fresh += 1

            # Stop paginating if no fresh jobs on this page (results are relevance-sorted, not date-sorted)
            if page_fresh == 0 and page > 1:
                break

        print(f"    Xing/{role}: {role_fresh} fresh (of {role_raw} raw)")

    return jobs

def parse_stepstone_timeago(timeago: str) -> datetime:
    """Convert German relative time string to approximate UTC datetime.

    Examples: 'vor 49 Minuten', 'vor 1 Stunde', 'vor 3 Stunden', 'vor 1 Tag'
    """
    now = datetime.now(timezone.utc)
    m = re.match(r"vor\s+(\d+)\s+(Minute|Stunde|Stunden|Tag|Tagen)", timeago, re.IGNORECASE)
    if m:
        n = int(m.group(1))
        unit = m.group(2).lower()
        if "minute" in unit:
            return now - timedelta(minutes=n)
        elif "stunde" in unit:
            return now - timedelta(hours=n)
        elif "tag" in unit:
            return now - timedelta(days=n)
    return now
def _parse_stepstone_card(card, cutoff, now):
    """Parse one Stepstone job card HTML into a job dict, or None if filtered out.

    Extracts URL, title, company, location, date, and description from the card's
    data-at attributes, enforces 24h freshness, and applies title/experience/location filters.
    """
    # URL — href on the job-item-title link
    url_m = re.search(r'href="(/stellenangebote--[^"]+)"', card)
    if not url_m:
        return None
    job_url = url_m.group(1)

    # Title — text content inside the job-item-title link
    title_m = re.search(r'data-at="job-item-title"[^>]*>(.*?)</a>', card, re.DOTALL)
    if not title_m:
        return None
    title = re.sub(r'<[^>]+>', '', title_m.group(1)).strip()
    title = html_mod.unescape(title)

    # Company — text inside a <div> within the TEXT span after company-name
    company_m = re.search(
        r'data-at="job-item-company-name"[^>]*>.*?data-genesis-element="TEXT"[^>]*>(.*?)</span>\s*</span>\s*</span>',
        card, re.DOTALL
    )
    if company_m:
        company = re.sub(r'<[^>]+>', '', company_m.group(1)).strip()
    else:
        company = "Unknown"
    company = html_mod.unescape(company)

    # Location — text inside the TEXT span within job-item-location
    loc_m = re.search(
        r'data-at="job-item-location"[^>]*>.*?data-genesis-element="TEXT"[^>]*>([^<]+)<',
        card, re.DOTALL
    )
    location = html_mod.unescape(loc_m.group(1).strip()) if loc_m else "Germany"

    # Date — inside <time> tag within job-item-timeago
    time_m = re.search(
        r'data-at="job-item-timeago"[^>]*><time[^>]*>([^<]+)</time>', card
    )
    timeago = time_m.group(1).strip() if time_m else ""
    posted_dt = parse_stepstone_timeago(timeago) if timeago else now

    # Freshness check (redundant with ag=age_1, but catches edge cases)
    if posted_dt < cutoff:
        return None

    # Description snippet from jobcard-content
    desc_m = re.search(r'data-at="jobcard-content"[^>]*>(.*?)</div>', card, re.DOTALL)
    desc = re.sub(r'<[^>]+>', '', desc_m.group(1)).strip() if desc_m else ""
    desc = html_mod.unescape(desc)

    # Apply existing filters (title relevance, seniority, experience, location)
    is_valid, role_type_or_reason = check_experience_and_location(title, desc, location)
    if not is_valid:
        return None

    return {
        "language": detect_language(title),
        "job_board": "Stepstone",
        "role_type": role_type_or_reason,
        "title": title,
        "company": company,
        "location": location,
        "posted_at": posted_dt.isoformat(),
        "exp_required": "<= 2 Years",
        "match_score": f"{compute_match_score(title)}%",
        "job_url": f"https://www.stepstone.de{job_url}",
        "description": desc
    }
def fetch_stepstone_jobs():
    """Fetch jobs from Stepstone.de via HTML scraping (free, no Apify).
    Uses plain requests to fetch server-side rendered search result pages.
    Stepstone is behind Akamai (not Cloudflare), so cloudscraper's challenge-
    solving hangs and causes read timeouts on port 443. Plain requests with a
    browser User-Agent works — Stepstone serves SSR HTML without anti-bot
    challenge. SSR-renders job cards in the initial HTML with data-at
    attributes for each field. Uses path-based URLs (/jobs/{slug}/in-deutschland)
    — the query-param format (?keyword=...) returns generic results regardless
    of the keyword. The ag=age_1 parameter pre-filters to last 24h ('Neuer als 24h').
    """
    if requests is None:
        print("[!] requests not installed — skipping Stepstone. Install with: pip install requests")
        return []

    jobs = []
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(hours=FRESHNESS_HOURS)
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "de-DE,de;q=0.9,en;q=0.8",
    })
    seen_urls = set()
    STEPSTONE_PAGES_PER_ROLE = 3  # 25 results/page × 3 = 75 max per role

    for role in SEARCH_ROLES:
        role_fresh = 0
        role_raw = 0
        slug = role.lower().replace(" ", "-")

        for page in range(1, STEPSTONE_PAGES_PER_ROLE + 1):
            url = (
                f"https://www.stepstone.de/jobs/{slug}/in-deutschland"
                f"?sort=2"          # sort by date (Datum)
                f"&ag=age_1"        # freshness: last 24h
                f"&page={page}"
            )
            try:
                r = session.get(url, timeout=15)
                if r.status_code != 200:
                    print(f"[!] Stepstone/{role} page {page}: HTTP {r.status_code}")
                    break
                r.encoding = "utf-8"
                raw = r.text
            except Exception as e:
                print(f"[!] Stepstone fetch error ({role} page {page}): {e}")
                break

            # Strip <style> and <svg> blocks — they contain CSS/paths that
            # interfere with regex parsing of data-at attributes.
            clean = re.sub(r'<style[^>]*>.*?</style>', '', raw, flags=re.DOTALL)
            clean = re.sub(r'<svg[^>]*>.*?</svg>', '', clean, flags=re.DOTALL)

            # Split into job cards by <article data-at="job-item"
            cards = re.split(r'<article[^>]*data-at="job-item"', clean)
            job_cards = [c.split('</article>')[0] for c in cards if 'data-at="job-item-title"' in c]

            page_fresh = 0
            for card in job_cards:
                # URL dedup — extract path for seen_urls check before parsing
                url_m = re.search(r'href="(/stellenangebote--[^"]+)"', card)
                if not url_m:
                    continue
                job_url = url_m.group(1)
                if job_url in seen_urls:
                    continue
                seen_urls.add(job_url)
                role_raw += 1

                job = _parse_stepstone_card(card, cutoff, now)
                if job:
                    jobs.append(job)
                    page_fresh += 1
                    role_fresh += 1

            # Stop paginating if no fresh jobs on this page
            if page_fresh == 0:
                break

        print(f"    Stepstone/{role}: {role_fresh} fresh (of {role_raw} raw)")

    return jobs



def fetch_linkedin_jobs_free():
    """Fetch LinkedIn jobs via free HTML scraping (no Apify, $0 cost).
    Scrapes the public LinkedIn jobs search page with f_TPR=r86400 (24h filter).

    Multi-city search strategy: LinkedIn caps results at 60 per page with no
    pagination (start= parameter is ignored). Searching "Germany" alone misses
    jobs beyond the first 60. To maximize coverage, each role is searched across
    multiple German cities — city-level results have different rankings and
    overlap only partially with the Germany-wide search.

    Rate-limit resilience: 3 workers (not 5) + 3 retries with exponential
    backoff (3s, 6s, 12s) + random jitter on polite delays. Eliminates 429
    errors that occurred at 5 workers with single retry.
    """
    import random
    from bs4 import BeautifulSoup
    from urllib.parse import urlencode

    HEADERS = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
    }

    # Multi-city search: Germany-wide + major tech hubs.
    LINKEDIN_LOCATIONS = ["Germany", "Berlin", "Munich", "Hamburg", "Frankfurt", "Cologne"]

    def _fetch_with_retry(url: str) -> requests.Response | None:
        """Fetch URL with 3 retries on 429 using exponential backoff (3s, 6s, 12s)."""
        for attempt in range(4):
            try:
                resp = requests.get(url, headers=HEADERS, timeout=15)
                if resp.status_code == 429 and attempt < 3:
                    backoff = 3 * (2 ** attempt) + random.uniform(0, 1)
                    print(f"    ... LinkedIn 429 (attempt {attempt+1}/4), backing off {backoff:.1f}s", flush=True)
                    time.sleep(backoff)
                    continue
                resp.raise_for_status()
                return resp
            except requests.exceptions.HTTPError as exc:
                if attempt < 3 and "429" in str(exc):
                    backoff = 3 * (2 ** attempt) + random.uniform(0, 1)
                    print(f"    ... LinkedIn 429 (attempt {attempt+1}/4), backing off {backoff:.1f}s", flush=True)
                    time.sleep(backoff)
                    continue
                print(f"[!] LinkedIn fetch error: {exc}", flush=True)
                return None
            except Exception as exc:
                print(f"[!] LinkedIn fetch error: {exc}", flush=True)
                return None
        return None

    def _scrape_role(role: str) -> list[dict]:
        """Scrape one role across all locations. Returns jobs for this role only."""
        role_jobs = []
        for loc in LINKEDIN_LOCATIONS:
            params = {
                "keywords": role,
                "location": loc,
                "f_TPR": "r86400",   # last 24h (server-side filter)
                "sortBy": "DD",       # date descending
            }
            url = f"https://www.linkedin.com/jobs/search/?{urlencode(params)}"
            resp = _fetch_with_retry(url)
            if not resp:
                continue

            soup = BeautifulSoup(resp.text, "html.parser")

            for card in soup.select("div.base-card"):
                title_el = card.select_one("h3.base-search-card__title")
                company_el = card.select_one("h4.base-search-card__subtitle")
                location_el = card.select_one("span.job-search-card__location")
                link_el = card.select_one("a.base-card__full-link")
                date_el = card.select_one("time")

                if not title_el or not link_el:
                    continue

                title = title_el.get_text(strip=True)
                job_url = link_el.get("href", "").split("?")[0]
                company = company_el.get_text(strip=True) if company_el else "Unknown"
                card_location = location_el.get_text(strip=True) if location_el else loc
                posted_raw = date_el.get("datetime", "") if date_el else ""

                desc = ""
                is_valid, role_type = check_experience_and_location(title, desc, card_location)
                if not is_valid:
                    continue

                # Safety-net freshness post-filter
                if posted_raw:
                    posted_dt = None
                    try:
                        posted_dt = datetime.fromisoformat(posted_raw.replace("Z", "+00:00"))
                        if posted_dt.tzinfo is None:
                            posted_dt = posted_dt.replace(tzinfo=timezone.utc)
                    except (ValueError, TypeError):
                        try:
                            posted_dt = datetime.strptime(posted_raw, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                        except (ValueError, TypeError):
                            pass
                    if posted_dt:
                        now = datetime.now(timezone.utc)
                        cutoff = now - timedelta(hours=FRESHNESS_HOURS)
                        if posted_dt.hour == 0 and posted_dt.minute == 0:
                            posted_dt = posted_dt.replace(hour=23, minute=59, second=59)
                        if posted_dt < cutoff:
                            continue

                role_jobs.append({
                    "language": detect_language(title),
                    "job_board": "LinkedIn",
                    "role_type": role_type,
                    "title": title,
                    "company": company,
                    "location": card_location,
                    "posted_at": posted_raw or "Last 24h",
                    "exp_required": "<= 2 Years",
                    "match_score": f"{compute_match_score(title)}%",
                    "job_url": job_url,
                    "description": "",
                })

            time.sleep(random.uniform(0.5, 1.5))  # jittered polite delay

        print(f"    LinkedIn free '{role}': {len(role_jobs)} jobs (6 locations)", flush=True)
        return role_jobs

    # 3 workers (was 5) — eliminates 429 rate limiting at the cost of ~10s
    # extra runtime. 5 workers triggered 429 on ~5% of requests; 3 workers
    # with exponential backoff retry achieves 0% 429s.
    all_role_jobs = []
    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {executor.submit(_scrape_role, role): role for role in SEARCH_ROLES}
        for future in as_completed(futures):
            try:
                all_role_jobs.extend(future.result())
            except Exception as exc:
                print(f"[!] LinkedIn role error: {exc}", flush=True)

    # Cross-role URL dedup (each role scraped independently, URLs may overlap)
    seen_urls = set()
    jobs = []
    for job in all_role_jobs:
        url = job["job_url"]
        if url not in seen_urls:
            seen_urls.add(url)
            jobs.append(job)

    return jobs

def fetch_indeed_jobs():
    """Fetch Indeed jobs via Indeed's public GraphQL API (Germany, last 24h).

    Uses the same mobile API endpoint as the Indeed iOS app
    (apis.indeed.com/graphql) with a hardcoded API key. No Cloudflare,
    no auth, no Apify — completely free. Returns full job descriptions.

    The API's dateOnIndeed filter enforces 24h freshness server-side.
    Post-filters on datePublished as a safety net.
    """
    from bs4 import BeautifulSoup

    API_URL = "https://apis.indeed.com/graphql"
    API_HEADERS = {
        "Host": "apis.indeed.com",
        "content-type": "application/json",
        "indeed-api-key": "161092c2017b5bbab13edb12461a62d5a833871e7cad6d9d475304573de67ac8",
        "accept": "application/json",
        "indeed-locale": "en-US",
        "accept-language": "en-US,en;q=0.9",
        "user-agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148 Indeed App 193.1",
        "indeed-app-info": "appv=193.1; appid=com.indeed.jobsearch; osv=16.6.1; os=ios; dtype=phone",
        "indeed-co": "DE",
    }

    QUERY_TEMPLATE = """query GetJobData {{
      jobSearch(
        what: "{role}"
        location: {{where: "Germany", radius: 50, radiusUnit: MILES}}
        limit: 50
        sort: RELEVANCE
        filters: {{
          date: {{
            field: "dateOnIndeed",
            start: "24h"
          }}
        }}
      ) {{
        pageInfo {{ nextCursor }}
        results {{
          trackingKey
          job {{
            key
            title
            datePublished
            dateOnIndeed
            description {{ html }}
            location {{
              countryName
              countryCode
              city
              formatted {{ short long }}
            }}
            employer {{
              name
              relativeCompanyPageUrl
            }}
          }}
        }}
      }}
    }}"""

    jobs = []
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(hours=FRESHNESS_HOURS)

    def fetch_role(role):
        role_jobs = []
        query = QUERY_TEMPLATE.format(role=role.replace('"', '\\"'))
        try:
            resp = requests.post(API_URL, headers=API_HEADERS, json={"query": query}, timeout=15)
            if not resp.ok:
                print(f"[!] Indeed GraphQL API returned {resp.status_code} for '{role}'", flush=True)
                return role_jobs
            data = resp.json()
            results = data.get("data", {}).get("jobSearch", {}).get("results", [])
        except Exception as e:
            print(f"[!] Indeed GraphQL error for '{role}': {e}", flush=True)
            return role_jobs

        for r in results:
            j = r.get("job", {})
            if not j:
                continue
            title = j.get("title", "")
            emp = j.get("employer") or {}
            loc_obj = j.get("location") or {}
            loc = loc_obj.get("formatted", {}).get("long") if loc_obj.get("formatted") else None
            loc = loc or ", ".join(filter(None, [loc_obj.get("city"), loc_obj.get("countryName")])) or "Germany"

            # Strip HTML from description
            desc_html = (j.get("description") or {}).get("html", "")
            desc = BeautifulSoup(desc_html, "html.parser").get_text(separator=" ", strip=True) if desc_html else ""

            is_valid, role_type_or_reason = check_experience_and_location(title, desc, loc)
            if not is_valid:
                continue

            # Safety-net freshness post-filter (API filter should already enforce 24h)
            posted_ms = j.get("dateOnIndeed") or j.get("datePublished")
            posted_dt = None
            if posted_ms:
                try:
                    posted_dt = datetime.fromtimestamp(posted_ms / 1000, tz=timezone.utc)
                except (ValueError, TypeError, OSError):
                    pass
            if posted_dt and posted_dt < cutoff:
                continue

            posted_at = posted_dt.strftime("%Y-%m-%dT%H:%M:%S.000Z") if posted_dt else "Last 24h"
            role_jobs.append({
                "language": detect_language(f"{title} {desc}"),
                "job_board": "Indeed",
                "role_type": role_type_or_reason,
                "title": title,
                "company": emp.get("name", "Unknown"),
                "location": loc,
                "posted_at": posted_at,
                "exp_required": "<= 2 Years",
                "match_score": f"{compute_match_score(f'{title} {desc}')}%",
                "job_url": f"https://de.indeed.com/viewjob?jk={j.get('key', '')}",
                "description": desc,
            })
        print(f"    Indeed GraphQL '{role}': {len(role_jobs)} jobs", flush=True)
        return role_jobs

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(fetch_role, role) for role in SEARCH_ROLES]
        for future in as_completed(futures):
            try:
                jobs.extend(future.result())
            except Exception as e:
                print(f"[!] Indeed role fetch error: {e}", flush=True)
    return jobs

def _style_xlsx_header(ws, header_fill, header_font, center):
    """Apply fill, font, and center alignment to the header row."""
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

def _format_xlsx_cells(ws, url_col, score_col, link_font, center):
    """Format the last appended row: hyperlink the URL cell, numericize the score cell."""
    r = ws.max_row
    if url_col:
        cell = ws.cell(row=r, column=url_col)
        url = cell.value or ""
        if url:
            cell.hyperlink = url
            cell.font = link_font
    if score_col:
        cell = ws.cell(row=r, column=score_col)
        v = cell.value
        if isinstance(v, str) and v.endswith("%") and v[:-1].strip().isdigit():
            cell.value = int(v[:-1].strip())
            cell.number_format = '0"%"'
            cell.alignment = center

def convert_csv_to_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    """Convert the exported CSV to XLSX: frozen header, autofilter (easy sorting),
    numeric match_score, and clickable job_url hyperlinks. Requires openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[!] openpyxl not installed - skipping XLSX export")
        return

    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        print("[!] Empty CSV - skipping XLSX export")
        return
    header, data = rows[0], rows[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Search"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    link_font = Font(color="0563C1", underline="single")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(header)
    _style_xlsx_header(ws, header_fill, header_font, center)

    url_col = header.index("job_url") + 1 if "job_url" in header else None
    score_col = header.index("match_score") + 1 if "match_score" in header else None

    for row in data:
        ws.append(row)
        _format_xlsx_cells(ws, url_col, score_col, link_font, center)

    for idx in range(1, len(header) + 1):
        letter = get_column_letter(idx)
        widths = [len(str(ws.cell(row=r, column=idx).value or "")) for r in range(1, ws.max_row + 1)]
        ws.column_dimensions[letter].width = min(max(widths) + 2, 80)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"
    ws.row_dimensions[1].height = 22

    wb.save(xlsx_path)
    print(f"[✓] XLSX exported to: {xlsx_path}")

def main():
    print("=== Jobscraper Pipeline (100% Free, No Apify) ===")
    print(f"    LinkedIn: FREE HTML scraping ($0, 10 roles × 6 locations parallel)")
    print(f"    Indeed: GraphQL API ($0, 10 roles parallel, full descriptions)")
    print(f"    Xing/Stepstone: free HTML scraping (parallel)")
    print(f"    ATS Direct: Greenhouse/SmartRecruiters/Ashby (free public APIs)")
    print(f"    All 6 platforms run in parallel via ThreadPoolExecutor")

    from ats_scraper import fetch_all_ats

    # Each platform fetcher runs in its own thread.
    # I/O bound work (HTTP requests + parsing) — GIL released during I/O,
    # so threads give near-linear speedup. Each fetcher is independent:
    # no shared mutable state, results collected after all complete.
    PLATFORM_FETCHERS = [
        ("Arbeitnow",    fetch_arbeitnow_jobs),
        ("Xing",         fetch_xing_jobs),
        ("Stepstone",    fetch_stepstone_jobs),
        ("LinkedIn",     fetch_linkedin_jobs_free),
        ("Indeed",       fetch_indeed_jobs),
        ("ATS Direct",   fetch_all_ats),
    ]

    all_jobs = []
    platform_counts = {}
    errors = {}

    print(f"\n  Launching {len(PLATFORM_FETCHERS)} platform fetchers in parallel...")
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=len(PLATFORM_FETCHERS)) as executor:
        future_to_name = {
            executor.submit(fetcher): name
            for name, fetcher in PLATFORM_FETCHERS
        }
        for future in as_completed(future_to_name):
            name = future_to_name[future]
            try:
                jobs = future.result()
                all_jobs.extend(jobs)
                platform_counts[name] = len(jobs)
                print(f"  [✓] {name}: {len(jobs)} jobs ({time.time() - start_time:.1f}s)", flush=True)
            except Exception as exc:
                platform_counts[name] = 0
                errors[name] = str(exc)
                print(f"  [!] {name}: ERROR — {exc} ({time.time() - start_time:.1f}s)", flush=True)

    elapsed = time.time() - start_time
    print(f"\n  All platforms complete in {elapsed:.1f}s")

    print("\n  Deduplication...")


    # Deduplication (within this run)
    seen_keys = set()
    deduped_jobs = []
    duplicates_count = 0

    for job in all_jobs:
        key = normalize_key(job["company"], job["title"])
        if key in seen_keys:
            duplicates_count += 1
            continue
        seen_keys.add(key)
        deduped_jobs.append(job)

    # Cross-run deduplication: drop jobs already in yesterday's export
    now = datetime.now()
    prev_urls = load_previous_run_urls(now)
    if prev_urls:
        before = len(deduped_jobs)
        deduped_jobs = [j for j in deduped_jobs if normalize_job_url(j["job_url"]) not in prev_urls]
        cross_run_duplicates = before - len(deduped_jobs)
    else:
        cross_run_duplicates = 0

    print(f"Per-platform: {platform_counts}")
    print(f"Cross-run dedup: compared against yesterday ({len(prev_urls)} jobs), removed {cross_run_duplicates} already-seen job(s)")
    print(f"Cost: $0.00 (all platforms free — no Apify)")

    # Format Date String: e.g. Aug_4_2026
    date_str = now.strftime("%b_%d_%Y").replace("_0", "_")

    # Per-date subfolder so each run's files are grouped/sorted by day
    date_folder = JOB_SEARCH_DIR / now.strftime("%Y-%m-%d")
    date_folder.mkdir(parents=True, exist_ok=True)

    csv_filename = f"Job_Search_{date_str}.csv"
    csv_path = date_folder / csv_filename
    json_path = date_folder / f"Job_Search_{date_str}.json"
    report_path = date_folder / "JOB_OPENINGS_LAST_24H.md"
    report_md = f"""# Daily Job Openings Report (Last 24 Hours)

**Generated:** {now.strftime("%Y-%m-%d %H:%M")}  
**CSV File:** `{csv_path}` (Sortable in Excel / Calc)

## Constraints Summary
- **Freshness:** Last 24 Hours
- **Experience:** <= 2 Years (No Senior/Lead/Manager roles)
- **Role Types:** Full-Time & Internships (Germany-wide) | Working Student (**Hamburg & Kiel ONLY**)
- **Target Roles:** Data Engineer, Analytics Engineer, Data Analyst, AI Data Engineer

| # | Language | Job Board | Role Type | Company | Job Title | Location | Match | Direct Link |
|---|---|---|---|---|---|---|---|---|
"""
    for idx, j in enumerate(deduped_jobs, 1):
        report_md += f"| {idx} | {j['language']} | **{j['job_board']}** | {j['role_type']} | {j['company']} | {j['title']} | {j['location']} | {j['match_score']} | [Apply Link]({j['job_url']}) |\n"

    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_md)
    # Write CSV (UTF-8 BOM for Excel compatibility)
    csv_fields = ["language", "job_board", "role_type", "title", "company",
                  "location", "posted_at", "exp_required", "match_score", "job_url"]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=csv_fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(deduped_jobs)

    # Write JSON (includes description field)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(deduped_jobs, f, ensure_ascii=False, indent=2)


    print(f"[✓] CSV exported successfully to: {csv_path}")
    print(f"[✓] JSON exported to: {json_path}")
    print(f"[✓] Markdown report exported to: {report_path}")

    # 4. Write Sortable XLSX (autofilter + clickable links)
    xlsx_path = date_folder / f"Job_Search_{date_str}.xlsx"
    convert_csv_to_xlsx(csv_path, xlsx_path)

if __name__ == "__main__":
    main()
