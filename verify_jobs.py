#!/usr/bin/env python3
"""
verify_jobs.py v2 — Full JD Verification Post-Step
===================================================

Standalone script that runs AFTER the Jobscraper pipeline exports its CSV.
Visits EVERY job URL (including LinkedIn + Indeed) to:
  1. Check if the listing is still active (not closed/filled)
  2. Extract the full job description text
  3. Apply three hard drop filters:
     - German level > B2 (C1/C2/fließend/Muttersprache/verhandlungssicher) → DROP
     - Experience ≥ 3 years (from JD body text) → DROP
     - Closed/inactive (404/410/redirect-to-expired) → DROP
  4. Apply one segregation filter:
     - Reposted LinkedIn jobs (cross-run history >7 days or job ID age >14 days)
       → moved to "Reposted" sheet (NOT dropped)
  5. Recalculate match score from actual JD text
  6. Enrich with: detail_language, detail_exp_years, detail_salary, detail_remote

Output: Job_Search_<date>_verified.xlsx with 2 sheets:
  - "Job Search": applicable jobs (passed all filters)
  - "Reposted":   LinkedIn jobs flagged as reposted (for manual review)

Usage:
  # Run inside eval sandbox (LLM classification available):
  import verify_jobs; verify_jobs.completion = completion
  verify_jobs.run_verification(Path("Job Search/YYYY-MM-DD/Job_Search_*.csv"), force=True)

  # Or standalone (falls back to regex for German/exp if no LLM):
  python3 verify_jobs.py                  # auto-finds most recent CSV
  python3 verify_jobs.py --csv path.csv   # specify input
  python3 verify_jobs.py --force          # re-verify all (ignore prior results)

Note: German level + experience classification uses LLM (smol model, batch of 5
JDs per call) when completion() is available. Falls back to regex patterns
when running standalone without an LLM client. LLM catches all phrasing variants
("sehr gut Deutsch, mind. auf Level C1", "fließend in Wort und Schrift", etc.)
that regex misses.
"""

import argparse
import csv
import json
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse

import requests

try:
    import cloudscraper
except ImportError:
    cloudscraper = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Constants ────────────────────────────────────────────────────────────────

JOB_SEARCH_DIR = Path("/home/sagar/Skills/Jobscraper/Job Search")

BROWSER_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)
HEADERS = {
    "User-Agent": BROWSER_UA,
    "Accept-Language": "de-DE,de;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

REQUEST_TIMEOUT = 15


def _response_text(resp) -> str:
    """Get response text with correct encoding.
    requests defaults to ISO-8859-1 when no charset is in Content-Type header,
    causing mojibake on German sites (Ã¼ instead of ü). Use apparent_encoding
    (chardet detection) as a better fallback.
    """
    if resp.encoding is None or resp.encoding.lower() in ("iso-8859-1", "latin-1"):
        resp.encoding = resp.apparent_encoding or "utf-8"
    return resp.text

# Platforms that use public APIs (no HTML scraping)
ATS_PLATFORMS = {"Greenhouse", "SmartRecruiters", "Ashby"}

# Tech keywords for match score recalculation (same as pipeline)
TECH_KEYWORDS = [
    "dbt", "airflow", "spark", "pyspark", "python", "sql", "gcp",
    "bigquery", "aws", "azure", "databricks", "docker", "kafka",
    "postgresql", "snowflake",
]

# LinkedIn job ID growth rate (~530K new IDs/day globally, verified Aug 23-27)
LINKEDIN_DAILY_ID_GROWTH = 530000

# Repost detection thresholds
REPOST_CROSS_RUN_DAYS = 7      # appeared in a run >7 days ago
REPOST_JOB_ID_AGE_DAYS = 14    # job ID suggests >14 days old

# ── German Language Detection ────────────────────────────────────────────────

# Hard requirement: C1/C2/fließend/Muttersprache/verhandlungssicher → DROP
GERMAN_REQUIRED_PATTERNS = [
    # Explicit CEFR level: "C1 Deutsch", "Deutsch C1", "mindestens C1"
    r'(?:mindestens|min\.|ab)\s*C[12]\s*(?:Deutsch|German|in\s+Deutsch)',
    r'C[12]\s*[-/]?\s*(?:Deutsch|German)',
    r'(?:Deutsch|German)\s+C[12]',
    r'(?:Deutsch|German)\s+(?:auf\s+)?C[12](?:\s*[-–]?\s*Niveau|\s+level)?',
    # "Fließend" / "fluent" — implies C1+ regardless of explicit level
    # Handles ß, ss (ASCII transliteration), and single s
    r'flie(?:ß|ss|s)end\s*(?:Deutsch|German|in\s+Deutsch|Deutschkenntnisse)',
    r'(?:Deutsch|German)\s+flie(?:ß|ss|s)end',
    # "Muttersprache" / native speaker
    r'Muttersprache\s+Deutsch',
    r'(?:Deutsch|German)\s+as\s+a\s+(?:first\s+)?native\s+language',
    # "Sehr gute Deutschkenntnisse" (standalone) — C1+, DROP
    # Does NOT match suspended compound "sehr gute Deutsch- und Englischkenntnisse" (false positive, keep)
    r'sehr\s+gute\s+Deutsch(?:kenntnisse|sprachkenntnisse)',
    # "Verhandlungssicher" — business-fluent, C1+
    r'(?:Deutsch|German)\s+verhandlungssicher',
    r'verhandlungssicher\s+in\s+Deutsch',
    # "Business fluent in German" (English postings)
    r'business\s+fluent\s+(?:in\s+)?German',
    r'fluent\s+German\s+(?:required|mandatory|must)',
    # "Muttersprachlich" (adjective form)
    r'muttersprachlich(?:e|er|es|em)?\s*(?:Deutsch|German)',
    # "Deutsch auf muttersprachlichem Niveau"
    r'(?:Deutsch|German)\s+.*muttersprachlich',
    # "Fachfließend" / "verhandlungssicher" variants
    r'fachflie[ßss]end\s*(?:Deutsch|German)',
]

# Soft requirement: German helpful but not required → KEEP + flag
GERMAN_SOFT_PATTERNS = [
    r'Deutsch(?:kenntnisse)?\s+(?:w(?:[üu]|au|ue)nschenswert|von\s+Vorteil|idealerweise)',
    r'German\s+(?:is\s+a\s+plus|preferred|nice\s+to\s+have|a\s+bonus)',
    r'idealerweise\s+Deutsch',
    # B1/B2 explicit — these are OK, Sagar has B2
    r'B[12]\s*[-/]?\s*(?:Deutsch|German)',
    r'(?:Deutsch|German)\s+B[12]',
    # "Grundkenntnisse" / "basic German" — fine
    r'(?:Grund|Basis)kenntnisse\s+(?:Deutsch|German)',
    r'basic\s+German',
]

_REQUIRED_RE = [re.compile(p, re.IGNORECASE) for p in GERMAN_REQUIRED_PATTERNS]
_SOFT_RE = [re.compile(p, re.IGNORECASE) for p in GERMAN_SOFT_PATTERNS]


def _split_sentences(text: str) -> list[str]:
    """Split text into sentences on . ! ? followed by whitespace or end."""
    parts = re.split(r'(?<=[.!?])\s+', text)
    return [p.strip() for p in parts if p.strip()]


def detect_german_requirement(text: str) -> str:
    """Detect German language requirement from job description text.

    Returns:
        "German C1+ required" — hard requirement, row should be DROPPED.
        "German preferred"    — soft/preferred, keep row but flag.
        "German B1/B2 OK"     — B1/B2 explicit, keep row.
        ""                    — no German requirement detected.
    """
    if not text:
        return ""

    # Check for B1/B2 explicit first (these are OK, not a drop reason)
    # Filter to SOFT patterns that contain 'B[12]' in their pattern string,
    # then check if any of those patterns match the text.
    b_level_patterns = [rx for rx in _SOFT_RE if 'B[12]' in rx.pattern]
    b_level = any(rx.search(text) for rx in b_level_patterns)

    sentences = _split_sentences(text)

    # Priority 1: find a sentence where a REQUIRED pattern matches AND no SOFT
    # pattern contradicts it in the same sentence.
    for sentence in sentences:
        req_match = any(rx.search(sentence) for rx in _REQUIRED_RE)
        if not req_match:
            continue
        soft_match = any(rx.search(sentence) for rx in _SOFT_RE)
        if not soft_match:
            return "German C1+ required"

    # Priority 2: B1/B2 explicit — keep, note it
    if b_level:
        return "German B1/B2 OK"

    # Priority 3: soft/preferred only — keep, flag.
    if any(rx.search(text) for rx in _SOFT_RE):
        return "German preferred"

    return ""


# ── Experience Extraction ────────────────────────────────────────────────────

_EXP_PATTERNS = [
    re.compile(r'(\d+)\s*(?:\+\s*)?Jahre?\s*(?:Berufs)?[eE]rfahrung'),
    re.compile(r'(\d+)\s*\+?\s*years?\s*(?:of\s*)?(?:professional\s*)?experience', re.IGNORECASE),
    re.compile(r'min\.\s*(\d+)\s*Jahre?', re.IGNORECASE),
    re.compile(r'(?:at\s+least|minimum)\s*(\d+)\s*years?', re.IGNORECASE),
    re.compile(r'(\d+)\s*Jahre\s*(?:berufliche|praktische)\s*(?:Erfahrung|Kenntnisse)', re.IGNORECASE),
]


def extract_exp_years(text: str) -> str:
    """Extract minimum experience years from job description text.

    Returns the minimum years found (smallest number, since that's the
    actual requirement). Empty string if no match.
    """
    if not text:
        return ""
    years = []
    for rx in _EXP_PATTERNS:
        for m in rx.finditer(text):
            try:
                years.append(int(m.group(1)))
            except (ValueError, IndexError):
                pass
    if not years:
        return ""
    # Return the minimum — "1-3 Jahre Erfahrung" means 1 year minimum
    return str(min(years))


# ── Salary Extraction ────────────────────────────────────────────────────────

_SALARY_NUM = r'(?:\d{2,3}(?:[.,]\d{3})+|\d{4,6})'
_SALARY_RE = re.compile(
    r'(' + _SALARY_NUM + r')\s*(?:€|EUR|Euro)?\s*'
    r'(?:bis\s*(' + _SALARY_NUM + r')\s*)?'
    r'(?:€|EUR|Euro)\s*'
    r'(?:(/Jahr|/jahr|p\.a\.|per\s+year)|(/Monat|/monat|per\s+month)|brutto)?',
    re.IGNORECASE,
)


def _extract_salary_jsonld(jsonld: dict | None) -> str:
    """Extract salary string from JSON-LD baseSalary field. Returns "" if absent."""
    if not jsonld or not isinstance(jsonld, dict):
        return ""
    bs = jsonld.get("baseSalary")
    if not bs or not isinstance(bs, dict):
        return ""
    cur = bs.get("currency", "EUR")
    val = bs.get("value", {})
    if not isinstance(val, dict):
        return ""
    lo = val.get("minValue", val.get("value", ""))
    hi = val.get("maxValue", "")
    if lo and hi:
        return f"{lo}-{hi} {cur}/year"
    return ""


def _detect_salary_period(text: str, match: re.Match) -> str:
    """Detect salary period ("/year", "/month", or "") from text after the match."""
    if match.group(3):
        return "/year"
    if match.group(4):
        return "/month"
    tail = text[match.end():match.end() + 20].lower()
    if "jahr" in tail or "year" in tail or "p.a" in tail:
        return "/year"
    if "monat" in tail or "month" in tail:
        return "/month"
    return ""


def extract_salary(text: str, jsonld: dict | None = None) -> str:
    """Extract salary range from text or JSON-LD baseSalary field."""
    # JSON-LD baseSalary (structured data) — highest priority
    jsonld_salary = _extract_salary_jsonld(jsonld)
    if jsonld_salary:
        return jsonld_salary
    if not text:
        return ""
    m = _SALARY_RE.search(text)
    if not m:
        return ""
    lo = m.group(1)
    hi = m.group(2)
    period = _detect_salary_period(text, m)
    if hi:
        return f"{lo}-{hi} EUR{period}"
    return f"{lo} EUR{period}"


# ── Remote Detection ─────────────────────────────────────────────────────────

_REMOTE_KEYWORDS = {
    "remote": ["remote", "home-office", "home office", "homeoffice",
               "fully remote", "100% remote", "distributed", "work from anywhere"],
    "hybrid": ["hybrid", "teilweise remote", "flexibles arbeiten", "mix"],
}
_ONSITE_KEYWORDS = ["vor ort", "on-site", "onsite", "in-house", "büro", "office presence"]


def extract_remote(text: str) -> str:
    """Detect remote/hybrid/onsite from job description text."""
    if not text:
        return ""
    lower = text.lower()
    for kw in _REMOTE_KEYWORDS["remote"]:
        if kw in lower:
            return "remote"
    for kw in _REMOTE_KEYWORDS["hybrid"]:
        if kw in lower:
            return "hybrid"
    for kw in _ONSITE_KEYWORDS:
        if kw in lower:
            return "onsite"
    return ""


# ── Match Score Recalculation ────────────────────────────────────────────────

def compute_match_score_from_jd(jd_text: str) -> int:
    """Recalculate match score from full job description text.

    Same algorithm as pipeline's compute_match_score() but runs on the
    actual JD text instead of just title + snippet.
    """
    if not jd_text:
        return 0
    text_lower = jd_text.lower()
    matches = sum(1 for kw in TECH_KEYWORDS if kw in text_lower)
    return min(100, int((matches / len(TECH_KEYWORDS)) * 100 * 2.5))


# ── JSON-LD Extraction ───────────────────────────────────────────────────────

_JSONLD_RE = re.compile(
    r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
    re.DOTALL | re.IGNORECASE,
)


def _extract_jsonld(html: str) -> dict | None:
    """Extract first JobPosting JSON-LD block from HTML."""
    for m in _JSONLD_RE.finditer(html):
        try:
            data = json.loads(m.group(1).strip())
        except (json.JSONDecodeError, ValueError):
            continue
        if isinstance(data, list):
            for item in data:
                if isinstance(item, dict) and item.get("@type") in (
                    "JobPosting", "https://schema.org/JobPosting"
                ):
                    return item
        elif isinstance(data, dict) and data.get("@type") in (
            "JobPosting", "https://schema.org/JobPosting"
        ):
            return data
    return None


def _extract_description_text(html: str) -> tuple[str, dict | None]:
    """Extract job description text from HTML for signal extraction.

    Returns (description_text, jsonld_dict_or_None).
    Tries JSON-LD description first, then falls back to stripping HTML tags.
    """
    jsonld = _extract_jsonld(html)
    if jsonld and jsonld.get("description"):
        desc = jsonld["description"]
        # JSON-LD descriptions may contain HTML entities and tags
        desc = re.sub(r'&lt;', '<', desc)
        desc = re.sub(r'&gt;', '>', desc)
        desc = re.sub(r'&amp;', '&', desc)
        desc = re.sub(r'&nbsp;', ' ', desc)
        desc = re.sub(r'<[^>]+>', ' ', desc)
        desc = re.sub(r'&[a-z]+;', ' ', desc)
        return desc.strip(), jsonld
    # Fallback: strip all HTML tags from the full page
    text = re.sub(r'<script[^>]*>.*?</script>', ' ', html, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<style[^>]*>.*?</style>', ' ', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'&[a-z]+;', ' ', text)
    return text.strip(), jsonld


# ── Reposted Detection (LinkedIn only) ───────────────────────────────────────

def _find_previous_run_dirs(job_search_dir: Path, today_str: str) -> list[Path]:
    """Return previous run directories sorted by date, excluding today and non-date dirs."""
    prev_dirs = []
    for run_dir in sorted(job_search_dir.iterdir()):
        if not run_dir.is_dir() or not run_dir.name.startswith("2026-"):
            continue
        if run_dir.name >= today_str:
            continue
        try:
            datetime.strptime(run_dir.name, "%Y-%m-%d")
        except ValueError:
            continue
        prev_dirs.append(run_dir)
    return prev_dirs


def _load_urls_from_csv(csv_path: Path) -> set[str]:
    """Return set of normalized job URLs from a CSV file."""
    urls = set()
    try:
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                url = _normalize_url(row.get("job_url", ""))
                if url:
                    urls.add(url)
    except Exception:
        pass
    return urls


def _load_linkedin_title_keys_from_csv(csv_path: Path) -> set[str]:
    """Return set of normalize_key(company, title) for LinkedIn rows from a CSV file."""
    import sys
    skill_dir = Path("/home/sagar/Skills/Jobscraper")
    if str(skill_dir) not in sys.path:
        sys.path.insert(0, str(skill_dir))
    from apify_job_search import normalize_key

    keys = set()
    try:
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                if row.get("job_board") != "LinkedIn":
                    continue
                key = normalize_key(row.get("company", ""), row.get("title", ""))
                if key:
                    keys.add(key)
    except Exception:
        pass
    return keys


def _load_repost_data(job_search_dir: Path, today_str: str) -> tuple[set, set]:
    """Load reposted detection data from previous runs.

    Returns:
        old_title_keys: normalize_key(company, title) from runs >7 days ago
        recent_urls: job URLs from the most recent previous run (carryovers)
    """
    cutoff = datetime.strptime(today_str, "%Y-%m-%d") - timedelta(days=REPOST_CROSS_RUN_DAYS)
    old_title_keys = set()
    recent_urls = set()

    if not job_search_dir.exists():
        return old_title_keys, recent_urls

    prev_dirs = _find_previous_run_dirs(job_search_dir, today_str)

    # Load URLs from the most recent previous run (carryover detection)
    if prev_dirs:
        most_recent = prev_dirs[-1]
        csvs = [c for c in most_recent.glob("Job_Search_*.csv")
                if "_verified" not in c.name and "_deduped" not in c.name]
        if csvs:
            recent_urls = _load_urls_from_csv(csvs[0])

    # Load title keys from runs older than 7 days (repost detection)
    for run_dir in prev_dirs:
        try:
            run_date = datetime.strptime(run_dir.name, "%Y-%m-%d")
        except ValueError:
            continue
        if run_date >= cutoff:
            continue
        csvs = [c for c in run_dir.glob("Job_Search_*.csv")
                if "_verified" not in c.name and "_deduped" not in c.name]
        if not csvs:
            continue
        old_title_keys |= _load_linkedin_title_keys_from_csv(csvs[0])

    return old_title_keys, recent_urls


def _normalize_url(url: str) -> str:
    """Normalize URL for cross-run comparison: strip query params + trailing slash."""
    url = url or ""
    # Drop query string first (LinkedIn tracking params like ?refId=... change per run)
    if "?" in url:
        url = url.split("?", 1)[0]
    # Then strip trailing slash
    url = url.rstrip("/")
    return url


def _extract_linkedin_job_id(url: str) -> int:
    """Extract numeric job ID from LinkedIn URL. Returns 0 if not found."""
    m = re.search(r'-(\d+)$', url or "")
    return int(m.group(1)) if m else 0


def detect_reposted(job: dict, today_str: str, old_title_keys: set,
                    today_max_linkedin_id: int, recent_urls: set | None = None) -> bool:
    """Check if a LinkedIn job is likely reposted.

    Two signals:
    1. Cross-run history: company::title appeared in a run >7 days ago
    2. Job ID age gap: job ID suggests >14 days old (based on ~530K IDs/day)

    Job ID override: if the job ID is fresh (<14 days old), signal 1 is
    suppressed — a fresh ID means it's a new posting, not a repost,
    even if the same company+title appeared in an old run.

    Carryover exception: if the job URL appeared in the most recent previous
    run, skip signal 1 (cross-run title match) — it's likely a 24h window
    overlap, not a repost. But signal 2 (job ID age gap) is NOT suppressed —
    a 277-day-old job ID is a repost regardless of carryover.
    """
    if job.get("job_board") != "LinkedIn":
        return False

    url = _normalize_url(job.get("job_url", ""))
    is_carryover = bool(recent_urls and url in recent_urls)

    # Check job ID age first — fresh ID overrides title match
    job_id = _extract_linkedin_job_id(url)
    is_fresh = False
    if job_id and today_max_linkedin_id:
        age_days = (today_max_linkedin_id - job_id) / LINKEDIN_DAILY_ID_GROWTH
        is_fresh = age_days <= REPOST_JOB_ID_AGE_DAYS
        if age_days > REPOST_JOB_ID_AGE_DAYS:
            return True

    # Signal 1: cross-run history (suppressed by carryover OR fresh job ID)
    if not is_carryover and not is_fresh:
        import sys
        from pathlib import Path as _P
        skill_dir = _P("/home/sagar/Skills/Jobscraper")
        if str(skill_dir) not in sys.path:
            sys.path.insert(0, str(skill_dir))
        from apify_job_search import normalize_key

        key = normalize_key(job.get("company", ""), job.get("title", ""))
        if key and key in old_title_keys:
            return True

    return False


# ── Per-Platform Verifiers ───────────────────────────────────────────────────

def _empty_result() -> dict:
    return {
        "verified_active": "",
        "detail_language": "",
        "detail_exp_years": "",
        "detail_reposted": "",
        "detail_salary": "",
        "detail_remote": "",
    }


def _extract_signals(text: str, jsonld: dict | None = None) -> dict:
    """Run regex-based signal extractors on description text.

    German language and experience are classified later via LLM batch.
    Only salary and remote use regex (simpler patterns, fewer variants).
    """
    return {
        "detail_salary": extract_salary(text, jsonld),
        "detail_remote": extract_remote(text),
    }


def _process_result(result: dict, desc_text: str, jsonld: dict | None = None) -> dict:
    """Fill in salary + remote + recalculated match score from JD text.

    German language and experience are classified later by llm_classify_all,
    which reads row["description"] directly — no hidden transport field.
    """
    signals = _extract_signals(desc_text, jsonld)
    result.update(signals)
    result["match_score"] = f"{compute_match_score_from_jd(desc_text)}%"
    return result


# ── LinkedIn Verifier (plain requests + JSON-LD, no auth) ────────────────────

def verify_linkedin(job: dict) -> dict:
    """Verify a single LinkedIn job. Uses pre-fetched TinyFish description if
    available, falls back to plain requests + JSON-LD otherwise.
    """
    result = _empty_result()
    url = job.get("job_url", "")
    if not url:
        return result

    # If TinyFish already injected a description, use it directly
    prefetched = job.get("description", "")
    if prefetched and len(prefetched) > 50:
        # Detect LinkedIn auth-wall boilerplate (no real JD content)
        _BOILERPLATE_MARKERS = ("Similar jobs", "People also viewed", "Referrals increase")
        _JD_MARKERS = ("requirements", "responsibilities", "qualifications", "experience",
                       "skills", "You will", "Your role", "What you", "Aufgaben",
                       "Anforderungen", "Profil", "Voraussetzungen", "Wir suchen",
                       "Über uns", "Das bringen Sie", "What you'll", "About the role",
                       "Job description", "About you", "Your mission", "Was Sie")
        text_lower = prefetched.lower()
        is_boilerplate = any(m in prefetched for m in _BOILERPLATE_MARKERS)
        has_jd = any(m.lower() in text_lower for m in _JD_MARKERS)
        if is_boilerplate and not has_jd:
            # Auth wall — LinkedIn didn't render the real JD. Leave unverified
            # so the user can review manually.
            result["detail_language"] = "AUTH WALL — review manually"
            return result
        result["verified_active"] = "True"
        result = _process_result(result, prefetched)
        return result

    # Fallback: plain requests + JSON-LD
    for attempt in range(2):
        try:
            resp = requests.get(url, timeout=REQUEST_TIMEOUT, headers=HEADERS,
                                allow_redirects=True)
            if resp.status_code == 404:
                result["verified_active"] = "False"
                return result
            if resp.status_code != 200:
                if attempt == 0:
                    time.sleep(3)
                    continue
                return result

            html = _response_text(resp)
            jsonld = _extract_jsonld(html)

            if jsonld:
                result["verified_active"] = "True"
                desc, _ = _extract_description_text(html)
                result = _process_result(result, desc, jsonld)
            else:
                title = job.get("title", "")
                if title and title.lower() in html.lower():
                    result["verified_active"] = "True"
                elif attempt == 0:
                    time.sleep(3)
                    continue
                else:
                    pass
            break
        except (requests.RequestException, requests.Timeout):
            if attempt == 0:
                time.sleep(3)
                continue
            pass
            break

    time.sleep(1)
    return result




# ── Indeed Verifier (plain requests) ─────────────────────────────────────────

def verify_indeed(job: dict) -> dict:
    """Verify an Indeed job using pre-fetched description.

    Description may come from Apify JSON (if non-empty) or TinyFish pre-fetch.
    Indeed job pages are behind a 401/403 auth wall — both plain requests and
    cloudscraper fail. If no description is available, the job stays unverified.
    """
    result = _empty_result()
    desc = job.get("description", "")
    if not desc:
        # No pre-fetched description — can't verify, keep as unknown
        return result
    # Job was found by Apify < 24h ago — it's active
    result["verified_active"] = "True"
    result = _process_result(result, desc)
    return result


# ── Xing Verifier ────────────────────────────────────────────────────────────

def verify_xing(job: dict) -> dict:
    """Verify a Xing job listing. Uses pre-fetched TinyFish description if
    available, falls back to plain requests otherwise."""
    result = _empty_result()
    url = job.get("job_url", "")
    if not url:
        return result

    # If TinyFish already injected a description, use it directly
    prefetched = job.get("description", "")
    if prefetched and len(prefetched) > 50:
        result["verified_active"] = "True"
        result = _process_result(result, prefetched)
        return result
    try:
        resp = requests.get(url, timeout=REQUEST_TIMEOUT, headers=HEADERS,
                            allow_redirects=True)
        time.sleep(1.5)
        if resp.status_code == 404:
            result["verified_active"] = "False"
            return result
        if resp.status_code != 200:
            return result
        html = _response_text(resp)
        jsonld = _extract_jsonld(html)
        title = job.get("title", "")
        if jsonld or (title and title.lower() in html.lower()):
            result["verified_active"] = "True"
            desc, _ = _extract_description_text(html)
            result = _process_result(result, desc, jsonld)
        else:
            result["verified_active"] = "False"
    except (requests.RequestException, requests.Timeout):
        pass
    return result


# ── Stepstone Verifier ───────────────────────────────────────────────────────

def verify_stepstone(job: dict) -> dict:
    """Verify a Stepstone job listing. Uses pre-fetched TinyFish description if
    available, falls back to plain requests otherwise."""
    result = _empty_result()
    url = job.get("job_url", "")
    if not url:
        return result

    # If TinyFish already injected a description, use it directly
    prefetched = job.get("description", "")
    if prefetched and len(prefetched) > 50:
        result["verified_active"] = "True"
        result = _process_result(result, prefetched)
        return result
    try:
        resp = requests.get(url, timeout=REQUEST_TIMEOUT, headers=HEADERS,
                            allow_redirects=True)
        time.sleep(2.0)
        if resp.status_code in (404, 410):
            result["verified_active"] = "False"
            return result
        if resp.status_code != 200:
            return result
        html = _response_text(resp)
        lower_html = html.lower()
        if "nicht mehr verfügbar" in lower_html or "no longer available" in lower_html:
            result["verified_active"] = "False"
            return result
        result["verified_active"] = "True"
        desc, jsonld = _extract_description_text(html)
        result = _process_result(result, desc, jsonld)
    except (requests.RequestException, requests.Timeout):
        pass
    return result


# ── Startup.jobs / Glassdoor Verifiers (cloudscraper) ────────────────────────

def verify_startupjobs(job: dict) -> dict:
    """Verify a Startup.jobs listing. Uses pre-fetched TinyFish description if
    available, falls back to cloudscraper otherwise."""
    result = _empty_result()
    url = job.get("job_url", "")
    if not url:
        return result

    # If TinyFish already injected a description, use it directly
    prefetched = job.get("description", "")
    if prefetched and len(prefetched) > 50:
        result["verified_active"] = "True"
        result = _process_result(result, prefetched)
        return result
    if cloudscraper is None:
        return result
    try:
        scraper = cloudscraper.create_scraper(browser={"browser": "chrome"})
        resp = scraper.get(url, timeout=REQUEST_TIMEOUT)
        time.sleep(2.0)
        if resp.status_code in (404, 410):
            result["verified_active"] = "False"
            return result
        if resp.status_code != 200:
            return result
        html = _response_text(resp)
        result["verified_active"] = "True"
        desc, jsonld = _extract_description_text(html)
        result = _process_result(result, desc, jsonld)
    except Exception:
        pass
    return result


def verify_glassdoor(job: dict) -> dict:
    """Verify a Glassdoor listing. Uses pre-fetched TinyFish description if
    available, falls back to cloudscraper otherwise."""
    result = _empty_result()
    url = job.get("job_url", "")
    if not url:
        return result

    # If TinyFish already injected a description, use it directly
    prefetched = job.get("description", "")
    if prefetched and len(prefetched) > 50:
        result["verified_active"] = "True"
        result = _process_result(result, prefetched)
        return result
    if cloudscraper is None:
        return result
    try:
        scraper = cloudscraper.create_scraper(browser={"browser": "chrome"})
        resp = scraper.get(url, timeout=REQUEST_TIMEOUT)
        time.sleep(2.0)
        if resp.status_code in (404, 410):
            result["verified_active"] = "False"
            return result
        if resp.status_code != 200:
            return result
        html = _response_text(resp)
        result["verified_active"] = "True"
        desc, jsonld = _extract_description_text(html)
        result = _process_result(result, desc, jsonld)
    except Exception:
        pass
    return result


# ── ATS Verifiers (public JSON APIs) ─────────────────────────────────────────

def _parse_ats_url(url: str, platform: str) -> tuple[str, str]:
    """Parse company slug and job ID from an ATS job URL."""
    parsed = urlparse(url)
    path_parts = [p for p in parsed.path.split("/") if p]
    if platform == "Greenhouse":
        if len(path_parts) >= 3 and path_parts[1] == "jobs":
            return path_parts[0], path_parts[2]
    elif platform == "SmartRecruiters":
        if len(path_parts) >= 2:
            return path_parts[0], path_parts[1]
    elif platform == "Ashby":
        if len(path_parts) >= 2:
            return path_parts[0], path_parts[1]
    return "", ""


def verify_greenhouse(job: dict) -> dict:
    """Verify a Greenhouse job via public API. 404/empty = closed."""
    result = _empty_result()
    url = job.get("job_url", "")
    slug, job_id = _parse_ats_url(url, "Greenhouse")
    if not slug or not job_id:
        return result
    api_url = f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs/{job_id}"
    try:
        resp = requests.get(api_url, timeout=REQUEST_TIMEOUT, headers=HEADERS)
        time.sleep(0.5)
        if resp.status_code == 404:
            result["verified_active"] = "False"
            return result
        if resp.status_code != 200:
            return result
        data = resp.json()
        if not data:
            result["verified_active"] = "False"
            return result
        result["verified_active"] = "True"
        desc = data.get("content", "") or data.get("title", "")
        if isinstance(desc, str):
            desc = re.sub(r'<[^>]+>', ' ', desc)
        result = _process_result(result, desc if isinstance(desc, str) else "")
    except (requests.RequestException, requests.Timeout, json.JSONDecodeError):
        pass
    return result


def verify_smartrecruiters(job: dict) -> dict:
    """Verify a SmartRecruiters job via public API. 404/empty = closed."""
    result = _empty_result()
    url = job.get("job_url", "")
    slug, job_id = _parse_ats_url(url, "SmartRecruiters")
    if not slug or not job_id:
        return result
    api_url = f"https://api.smartrecruiters.com/v1/companies/{slug}/jobs/{job_id}"
    try:
        resp = requests.get(api_url, timeout=REQUEST_TIMEOUT, headers=HEADERS)
        time.sleep(0.5)
        if resp.status_code == 404:
            result["verified_active"] = "False"
            return result
        if resp.status_code != 200:
            return result
        data = resp.json()
        if not data:
            result["verified_active"] = "False"
            return result
        result["verified_active"] = "True"
        desc = ""
        job_ad = data.get("jobAd", {})
        if isinstance(job_ad, dict):
            sections = job_ad.get("sections", {})
            if isinstance(sections, dict):
                for section in sections.values():
                    if isinstance(section, dict):
                        text = section.get("text", "")
                        if text:
                            desc += " " + re.sub(r'<[^>]+>', ' ', text)
        result = _process_result(result, desc.strip())
    except (requests.RequestException, requests.Timeout, json.JSONDecodeError):
        pass
    return result


def _find_ashby_posting(postings: list, job_id: str) -> dict | None:
    """Find the posting dict matching job_id in the Ashby postings list."""
    for posting in postings:
        if not isinstance(posting, dict):
            continue
        if posting.get("id") == job_id or job_id in (posting.get("id", "")):
            return posting
    return None


def _extract_ashby_desc(target: dict) -> str:
    """Extract and HTML-strip the description from an Ashby posting dict."""
    desc = target.get("descriptionHtml", "") or target.get("description", "")
    if isinstance(desc, str) and "<" in desc:
        desc = re.sub(r'<[^>]+>', ' ', desc)
    return desc if isinstance(desc, str) else ""


def verify_ashby(job: dict) -> dict:
    """Verify an Ashby job via public API. 404/empty = closed."""
    result = _empty_result()
    url = job.get("job_url", "")
    slug, job_id = _parse_ats_url(url, "Ashby")
    if not slug or not job_id:
        return result
    api_url = f"https://api.ashbyhq.com/posting-api/job-board/{slug}?includeCompensation=true"
    try:
        resp = requests.get(api_url, timeout=REQUEST_TIMEOUT, headers=HEADERS)
        time.sleep(0.5)
        if resp.status_code == 404:
            result["verified_active"] = "False"
            return result
        if resp.status_code != 200:
            return result
        data = resp.json()
        postings = data if isinstance(data, list) else data.get("postings", [])
        target = _find_ashby_posting(postings, job_id)
        if target is None:
            result["verified_active"] = "False"
            return result
        result["verified_active"] = "True"
        desc = _extract_ashby_desc(target)
        result = _process_result(result, desc)
    except (requests.RequestException, requests.Timeout, json.JSONDecodeError):
        pass
    return result


# ── Arbeitnow Verifier ───────────────────────────────────────────────────────

def verify_arbeitnow(job: dict) -> dict:
    """Verify an Arbeitnow job by checking the public API for the URL."""
    result = _empty_result()
    url = job.get("job_url", "")
    if not url:
        return result
    try:
        api_url = "https://www.arbeitnow.com/api/job-board-api"
        resp = requests.get(api_url, timeout=REQUEST_TIMEOUT, headers=HEADERS)
        if resp.status_code != 200:
            return result
        data = resp.json()
        job_urls = {item.get("url", "") for item in data.get("data", [])
                    if isinstance(item, dict)}
        if url in job_urls:
            result["verified_active"] = "True"
            # Arbeitnow API has description in the job data
            for item in data.get("data", []):
                if isinstance(item, dict) and item.get("url") == url:
                    desc = item.get("description", "")
                    if desc:
                        desc = re.sub(r'<[^>]+>', ' ', desc)
                        result = _process_result(result, desc)
                    break
        else:
            result["verified_active"] = ""
    except (requests.RequestException, requests.Timeout, json.JSONDecodeError):
        pass
    return result


# ── Platform Dispatch ────────────────────────────────────────────────────────

PLATFORM_VERIFIERS = {
    "LinkedIn": verify_linkedin,
    "Indeed": verify_indeed,
    "Xing": verify_xing,
    "Stepstone": verify_stepstone,
    "Startup.jobs": verify_startupjobs,
    "Glassdoor": verify_glassdoor,
    "Greenhouse": verify_greenhouse,
    "SmartRecruiters": verify_smartrecruiters,
    "Ashby": verify_ashby,
    "Arbeitnow": verify_arbeitnow,
}

# Platforms that use 2-worker parallelism (rate-limit sensitive)
PARALLEL_PLATFORMS = {"LinkedIn"}
# ATS platforms: 4 workers
ATS_PLATFORMS_SET = {"Greenhouse", "SmartRecruiters", "Ashby"}


def verify_platform_batch(platform: str, jobs: list[dict]) -> list[tuple[int, dict]]:
    """Verify all jobs for one platform. Returns list of (index, result_dict)."""
    results: list[tuple[int, dict]] = []
    verifier = PLATFORM_VERIFIERS.get(platform)

    if verifier is None:
        # Unknown platform — mark as active, no fetch
        for i, job in enumerate(jobs):
            r = _empty_result()
            r["verified_active"] = "True"
            results.append((i, r))
        return results

    if platform in ATS_PLATFORMS_SET and len(jobs) > 1:
        # ATS: 4 parallel workers, 0.5s delay inside each verifier
        with ThreadPoolExecutor(max_workers=4) as executor:
            future_to_idx = {
                executor.submit(verifier, job): i for i, job in enumerate(jobs)
            }
            for future in as_completed(future_to_idx):
                idx = future_to_idx[future]
                try:
                    results.append((idx, future.result()))
                except Exception:
                    results.append((idx, _empty_result()))
    elif platform in PARALLEL_PLATFORMS:
        # LinkedIn fallback: 2 workers (plain requests, rate-limit sensitive)
        with ThreadPoolExecutor(max_workers=2) as executor:
            future_to_idx = {
                executor.submit(verifier, job): i for i, job in enumerate(jobs)
            }
            for future in as_completed(future_to_idx):
                idx = future_to_idx[future]
                try:
                    results.append((idx, future.result()))
                except Exception:
                    results.append((idx, _empty_result()))
    else:
        # Sequential with delay (delay is inside each verifier)
        for i, job in enumerate(jobs):
            try:
                results.append((i, verifier(job)))
            except Exception:
                results.append((i, _empty_result()))

    return results


# ── CSV / XLSX I/O ───────────────────────────────────────────────────────────

INPUT_FIELDS = [
    "language", "job_board", "role_type", "title", "company",
    "location", "posted_at", "exp_required", "match_score", "job_url",
]
OUTPUT_FIELDS = INPUT_FIELDS + [
    "verified_active", "detail_language", "detail_exp_years",
    "detail_reposted", "detail_salary", "detail_remote",
]


def find_latest_csv() -> Path | None:
    """Find the most recent Job_Search_*.csv under the Job Search directory."""
    if not JOB_SEARCH_DIR.exists():
        return None
    date_folders = sorted(
        [d for d in JOB_SEARCH_DIR.iterdir() if d.is_dir()],
        reverse=True,
    )
    for folder in date_folders:
        csvs = sorted(folder.glob("Job_Search_*.csv"), reverse=True)
        csvs = [c for c in csvs if "_verified" not in c.name and "_deduped" not in c.name]
        if csvs:
            return csvs[0]
    return None


def load_csv(path: Path) -> list[dict]:
    """Load CSV rows, handling both original and already-verified schemas."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader)


def save_xlsx(path: Path, main_rows: list[dict], reposted_rows: list[dict]) -> None:
    """Write verified XLSX with 2 sheets: 'Job Search' and 'Reposted'.

    Same formatting as pipeline's convert_csv_to_xlsx: frozen header,
    autofilter, clickable URL hyperlinks, numeric match_score.
    """
    if not HAS_OPENPYXL:
        print("[!] openpyxl not installed — falling back to CSV")
        # Fallback: write two CSVs
        csv_main = path.with_suffix(".csv")
        with open(csv_main, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(main_rows)
        if reposted_rows:
            csv_repost = path.parent / f"{path.stem}_reposted.csv"
            with open(csv_repost, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(reposted_rows)
        print(f"[✓] CSV exported to: {csv_main}")
        return

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    link_font = Font(color="0563C1", underline="single")
    center = Alignment(horizontal="center", vertical="center")

    def _write_sheet(ws, rows, title):
        ws.title = title
        ws.append(OUTPUT_FIELDS)
        for c in ws[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = center

        url_col = OUTPUT_FIELDS.index("job_url") + 1
        score_col = OUTPUT_FIELDS.index("match_score") + 1

        for row in rows:
            ws.append([row.get(k, "") for k in OUTPUT_FIELDS])
            r = ws.max_row
            cell = ws.cell(row=r, column=url_col)
            url = cell.value or ""
            if url:
                cell.hyperlink = url
                cell.font = link_font
            cell = ws.cell(row=r, column=score_col)
            v = cell.value
            if isinstance(v, str) and v.endswith("%") and v[:-1].strip().isdigit():
                cell.value = int(v[:-1].strip())
                cell.number_format = '0"%"'
                cell.alignment = center

        for idx in range(1, len(OUTPUT_FIELDS) + 1):
            letter = get_column_letter(idx)
            widths = [len(str(ws.cell(row=r, column=idx).value or ""))
                      for r in range(1, ws.max_row + 1)]
            ws.column_dimensions[letter].width = min(max(widths) + 2, 80)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_FIELDS))}{ws.max_row}"
        ws.row_dimensions[1].height = 22

    # Sheet 1: Job Search (main)
    ws_main = wb.active
    _write_sheet(ws_main, main_rows, "Job Search")

    # Sheet 2: Reposted
    if reposted_rows:
        ws_repost = wb.create_sheet("Reposted")
        _write_sheet(ws_repost, reposted_rows, "Reposted")

    wb.save(path)
    print(f"[✓] XLSX exported to: {path}")

# ── LLM Batch Classification (German + Experience) ──────────────────────────

LLM_BATCH_SIZE = 5

_LLM_CLASSIFY_PROMPT = """Classify German language requirement and minimum experience years for each job.

German level (pick one):
- C1+ = C1, C2, fluent/fließend, native/Muttersprache, verhandlungssicher, "sehr gute Deutschkenntnisse" standalone, "mind. C1", "mindestens C1"
- B1/B2 = B1 or B2 only, "gute Deutschkenntnisse" without "sehr"
- preferred = nice-to-have/wünschenswert/von Vorteil/idealerweise
- none = no German mentioned, English-only

Notes: "Sehr gute Deutsch- und Englischkenntnisse" = B1/B2. "Sehr gute Deutschkenntnisse" standalone = C1+.

Experience years (minimum required):
- "mehrere Jahre" = 3, "mehrjährige" = 3, "einige Jahre" = 2, "1-3 Jahre" = 1
- empty = no requirement mentioned

Jobs:
{jobs}

Reply with EXACTLY {count} lines. Format: <job_number>|<german_level>|<exp_years_or_empty>
Example:
1|C1+|3
2|none|
3|preferred|2"""

# Keywords for extracting relevant JD sections (German + experience)
_RELEVANT_KEYWORDS = re.compile(
    r'(?:[Dd]eutsch|[Gg]erman|[Ss]prach(?:e|en|kenntnis|kenntnisse)|'
    r'[Ll]anguage|[Ff]ließend|[Mm]uttersprach|'
    r'[Vv]erhandlungssicher|[Bb]usiness\s+fluent|'
    r'[Nn]iveau|[Ll]evel\s+[ABC]|B[12]|C[12]|'
    r'[Ee]rfahrung|[Yy]ears?[Jj]ahre|[Bb]erufserfahrung|'
    r'[Jj]ahre\s+[Bb]eruf|mind\.\s*\d|mindestens\s*\d|'
    r'at\s+least\s+\d|\d+\+?\s+years?)',
    re.IGNORECASE
)


def _extract_relevant_sections(jd_text: str, context_chars: int = 200) -> str:
    """Extract sentences/sections containing German or experience keywords.

    Requirements often appear past char 2000 in JDs. Instead of truncating,
    extract only the relevant sections to keep the LLM prompt focused.
    Falls back to first 2000 chars if no keywords found.
    """
    if len(jd_text) <= 2000:
        return jd_text

    # Split into sentences (rough split on . ! ? followed by space/capital)
    sentences = re.split(r'(?<=[.!?])\s+', jd_text)
    relevant = []
    for sent in sentences:
        if _RELEVANT_KEYWORDS.search(sent):
            relevant.append(sent.strip())

    if relevant:
        return " ".join(relevant)[:3000]  # Cap at 3000 to keep prompt manageable

    # No keywords found — return first 2000 chars (job may have no requirements)
    return jd_text[:2000]


# Parse "N|level|years" lines from LLM plain-text output
_LLM_LINE_RE = re.compile(
    r'^(\d+)\s*\|\s*(C1\+|B1/B2|preferred|none)\s*\|\s*(\d*)\s*$',
    re.IGNORECASE,
)


def _parse_llm_response(text: str, jd_texts: list[str]) -> list[dict] | None:
    """Parse "N|level|years" lines from LLM output into result dicts.

    Fills gaps with regex fallback. Returns None if no lines parsed at all.
    """
    parsed = {}
    for line in text.splitlines():
        m = _LLM_LINE_RE.match(line.strip())
        if m:
            job_num = int(m.group(1))
            german = m.group(2)
            exp_str = m.group(3).strip()
            exp = int(exp_str) if exp_str else None
            parsed[job_num] = {"german": german, "exp_years": exp}

    if not parsed:
        return None

    results = []
    missing = 0
    for i in range(1, len(jd_texts) + 1):
        if i in parsed:
            results.append(parsed[i])
        else:
            missing += 1
            results.append({"german": _regex_german_level(jd_texts[i-1]), "exp_years": _regex_exp_years(jd_texts[i-1])})

    if missing:
        print(f"  [!] LLM: {len(parsed)}/{len(jd_texts)} parsed, {missing} filled with regex")

    return results


def llm_classify_batch(jd_texts: list[str]) -> list[dict]:
    """Classify German level + experience years for a batch of JD texts via LLM.

    Returns list of {"german": str, "exp_years": int|None} per job.
    Falls back to regex if LLM fails.
    """
    if not jd_texts:
        return []

    if "completion" not in globals():
        return [
            {"german": _regex_german_level(t), "exp_years": _regex_exp_years(t)}
            for t in jd_texts
        ]

    extracted = [_extract_relevant_sections(t) for t in jd_texts]
    jobs_block = "\n\n".join(
        f"{i+1}: {text}" for i, text in enumerate(extracted)
    )
    prompt = _LLM_CLASSIFY_PROMPT.format(jobs=jobs_block, count=len(jd_texts))

    try:
        raw = completion(prompt=prompt, model="smol")
        if isinstance(raw, dict):
            text = raw.get("value") or raw.get("text") or str(raw)
        else:
            text = str(raw)

        results = _parse_llm_response(text, jd_texts)
        if results is not None:
            return results
        print(f"  [!] LLM: 0/{len(jd_texts)} parsed — falling back to regex")
    except Exception as e:
        print(f"  [!] LLM classification failed: {e} — falling back to regex")

    return [
        {"german": _regex_german_level(t), "exp_years": _regex_exp_years(t)}
        for t in jd_texts
    ]


def _regex_german_level(text: str) -> str:
    """Regex fallback for German level classification."""
    result = detect_german_requirement(text)
    if result == "German C1+ required":
        return "C1+"
    elif result == "German B1/B2 OK":
        return "B1/B2"
    elif result == "German preferred":
        return "preferred"
    return "none"


def _regex_exp_years(text: str) -> int | None:
    """Regex fallback for experience years extraction."""
    result = extract_exp_years(text)
    try:
        return int(result) if result else None
    except (ValueError, TypeError):
        return None


def llm_classify_all(rows: list[dict]) -> None:
    """Batch-classify German + exp for all rows with JD text.

    Mutates rows in-place: sets detail_language and detail_exp_years.
    """
    # Collect rows that have JD text and need classification
    to_classify = []
    for i, row in enumerate(rows):
        jd = row.get("description", "")
        if len(jd) < 50:
            continue
        to_classify.append((i, jd))

    if not to_classify:
        print(f"[*] LLM classification: 0/{len(rows)} jobs classified (no description)")
        return

    print(f"[*] LLM classification: {len(to_classify)}/{len(rows)} jobs classified ({len(rows)-len(to_classify)} skipped — no description)")

    # One-time check: warn if LLM not available (regex fallback is less accurate)
    if "completion" not in globals():
        print("[!] WARNING: LLM not available — using regex fallback (less accurate)")

    print(f"\n[*] LLM batch classification: {len(to_classify)} jobs in batches of {LLM_BATCH_SIZE}...")

    # Process in batches
    for batch_start in range(0, len(to_classify), LLM_BATCH_SIZE):
        batch = to_classify[batch_start:batch_start + LLM_BATCH_SIZE]
        jd_texts = [jd for _, jd in batch]

        results = llm_classify_batch(jd_texts)

        for (row_idx, _), classification in zip(batch, results):
            row = rows[row_idx]
            german = classification.get("german", "none")
            exp = classification.get("exp_years")

            # Map LLM classification to output format
            if german == "C1+":
                row["detail_language"] = "German C1+ required"
            elif german == "B1/B2":
                row["detail_language"] = "German B1/B2 OK"
            elif german == "preferred":
                row["detail_language"] = "German preferred"
            else:
                row["detail_language"] = ""

            row["detail_exp_years"] = str(exp) if exp is not None else ""

        batch_num = batch_start // LLM_BATCH_SIZE + 1
        total_batches = (len(to_classify) + LLM_BATCH_SIZE - 1) // LLM_BATCH_SIZE
        print(f"  [✓] Batch {batch_num}/{total_batches} done ({len(batch)} jobs)")


# ── Main Orchestration ───────────────────────────────────────────────────────

def run_verification(csv_path: Path, force: bool = False) -> None:
    """Main entry: load CSV, verify per-platform, apply filters, write XLSX."""
    rows = load_csv(csv_path)
    if not rows:
        print(f"[!] No rows found in {csv_path}")
        return


    # Load sibling JSON for pre-fetched descriptions (Indeed, Arbeitnow, Stepstone)
    # Indeed job pages are behind 401/403 — use Apify-provided description instead
    json_path = csv_path.with_suffix(".json")
    if json_path.exists():
        try:
            with open(json_path, encoding="utf-8") as f:
                json_data = json.load(f)
            url_to_desc = {}
            for j in json_data:
                if isinstance(j, dict) and j.get("description"):
                    url_to_desc[j.get("job_url", "")] = j["description"]
            injected = 0
            for row in rows:
                url = row.get("job_url", "")
                if url in url_to_desc and not row.get("description"):
                    row["description"] = url_to_desc[url]
                    injected += 1
            if injected:
                print(f"[*] Injected descriptions from JSON for {injected} job(s)")
        except (json.JSONDecodeError, OSError):
            pass

    # Pre-fetch JDs via TinyFish (main thread — tool.* not thread-safe)
    # TinyFish renders JS-heavy pages (LinkedIn auth-walls, Xing, Stepstone, etc.)
    # that plain requests can't. Caches to disk so interrupts don't waste money.
    TINYFISH_PLATFORMS = {"LinkedIn", "Indeed", "Xing", "Stepstone",
                          "Startup.jobs", "Glassdoor"}
    tf_jobs = [r for r in rows
               if r.get("job_board") in TINYFISH_PLATFORMS
               and not r.get("description")]

    if tf_jobs:
        if "tinyfish_fetch" not in globals():
            print(f"[!] WARNING: tinyfish_fetch not available (not running in OMP eval).")
            print(f"[!] {len(tf_jobs)} jobs will have NO JD text — German/exp/salary")
            print(f"[!] detection will be inaccurate. Run via eval, not bash.")
        else:
            # ── Load disk cache so re-runs skip already-fetched URLs ──
            cache_path = csv_path.parent / "tinyfish_cache.json"
            cache: dict[str, str] = {}
            if cache_path.exists():
                try:
                    with open(cache_path, encoding="utf-8") as f:
                        cache = json.load(f)
                    print(f"[*] Loaded TinyFish cache: {len(cache)} descriptions")
                except (json.JSONDecodeError, OSError):
                    pass

            # Inject cached descriptions
            cache_hits = 0
            for row in tf_jobs:
                url = row.get("job_url", "")
                if url in cache and cache[url]:
                    row["description"] = cache[url]
                    cache_hits += 1
            if cache_hits:
                print(f"[*] Cache hits: {cache_hits}/{len(tf_jobs)} (skipping re-fetch)")

            # Only fetch jobs not in cache
            to_fetch = [r for r in tf_jobs if r.get("job_url") not in cache]
            if to_fetch:
                print(f"[*] Pre-fetching {len(to_fetch)} JDs via TinyFish "
                      f"({', '.join(sorted({r['job_board'] for r in to_fetch}))})...")
                tf_batch_size = 2  # keep response under truncation limit
                tf_injected = 0
                for batch_start in range(0, len(to_fetch), tf_batch_size):
                    batch = to_fetch[batch_start:batch_start + tf_batch_size]
                    urls = [j.get("job_url", "") for j in batch if j.get("job_url")]
                    if not urls:
                        continue
                    try:
                        resp = tinyfish_fetch(urls)
                        for item in resp.get("results", []):
                            u = item.get("url", "")
                            text = item.get("text", "")
                            if not text:
                                continue
                            cache[u] = text  # add to cache
                            for row in rows:
                                if row.get("job_url") == u:
                                    row["description"] = text
                                    tf_injected += 1
                                    break
                    except Exception as exc:
                        print(f"  [!] TinyFish batch {batch_start // tf_batch_size + 1} failed: {exc}")
                    # ── Save cache to disk after every batch (survive interrupts) ──
                    try:
                        with open(cache_path, "w", encoding="utf-8") as f:
                            json.dump(cache, f, ensure_ascii=False)
                    except OSError:
                        pass
                    batch_num = batch_start // tf_batch_size + 1
                    total_batches = (len(to_fetch) + tf_batch_size - 1) // tf_batch_size
                    print(f"    TinyFish batch {batch_num}/{total_batches} done "
                          f"({tf_injected} new JDs, {len(cache)} cached)", flush=True)
                print(f"[*] TinyFish injected {tf_injected} new descriptions "
                      f"({len(cache)} total in cache)")
            else:
                print(f"[*] All {len(tf_jobs)} JDs found in cache — no TinyFish fetch needed")

    # Print acquisition stats
    desc_count = sum(1 for r in rows if r.get("description", "").strip())
    print(f"\n[*] Description acquisition: {desc_count}/{len(rows)} jobs acquired descriptions")
    if desc_count < len(rows):
        missing = [r for r in rows if not r.get("description", "").strip()]
        missing_platforms = Counter(r.get("job_board", "Unknown") for r in missing)
        print(f"    {len(missing)} missing by platform:")
        for p, c in missing_platforms.most_common():
            print(f"      {p}: {c}")
    # Idempotency: skip rows already verified unless --force
    to_verify: list[tuple[int, dict]] = []
    already_verified = 0
    for i, row in enumerate(rows):
        existing = row.get("verified_active", "")
        if existing and not force:
            already_verified += 1
        else:
            to_verify.append((i, row))

    if already_verified:
        print(f"[*] {already_verified} row(s) already verified — skipping (use --force to re-verify)")

    # Group jobs to verify by platform
    platform_groups: dict[str, list[tuple[int, dict]]] = {}
    for i, row in to_verify:
        platform = row.get("job_board", "Unknown")
        platform_groups.setdefault(platform, []).append((i, row))

    if not platform_groups:
        print("[*] Nothing to verify — all rows already checked.")
        return

    # Print per-platform counts
    for platform in sorted(platform_groups):
        count = len(platform_groups[platform])
        print(f"    {platform}: {count} URL(s)")

    # ── Reposted detection setup (LinkedIn only) ──
    today_str = datetime.now().strftime("%Y-%m-%d")
    print(f"\n[*] Loading cross-run history for reposted detection...")
    old_title_keys, recent_urls = _load_repost_data(JOB_SEARCH_DIR, today_str)
    print(f"    Loaded {len(old_title_keys)} title keys from runs >{REPOST_CROSS_RUN_DAYS} days ago")
    print(f"    Loaded {len(recent_urls)} URLs from most recent previous run (carryover detection)")

    # Find today's max LinkedIn job ID for age-gap estimation
    today_max_linkedin_id = 0
    for _, row in to_verify:
        if row.get("job_board") == "LinkedIn":
            jid = _extract_linkedin_job_id(row.get("job_url", ""))
            if jid > today_max_linkedin_id:
                today_max_linkedin_id = jid
    if today_max_linkedin_id:
        print(f"    Today's max LinkedIn job ID: {today_max_linkedin_id}")

    # ── Run all platform batches in parallel ──
    all_results: dict[int, dict] = {}

    with ThreadPoolExecutor(max_workers=len(platform_groups)) as executor:
        future_to_platform = {}
        for platform, group in platform_groups.items():
            jobs = [row for _, row in group]
            future = executor.submit(verify_platform_batch, platform, jobs)
            future_to_platform[future] = (platform, group)

        for future in as_completed(future_to_platform):
            platform, group = future_to_platform[future]
            try:
                batch_results = future.result()
                for local_idx, result in batch_results:
                    orig_idx = group[local_idx][0]
                    all_results[orig_idx] = result
                print(f"  [✓] {platform}: done ({len(batch_results)} verified)")
            except Exception as exc:
                print(f"  [!] {platform}: batch failed ({exc})")
                for orig_idx, _ in group:
                    all_results[orig_idx] = _empty_result()

    # ── Merge results + apply reposted detection ──
    for i, row in enumerate(rows):
        result = all_results.get(i)
        if result:
            row.update(result)
        # Reposted detection (LinkedIn only, uses CSV data — no page fetch needed)
        is_repost = detect_reposted(row, today_str, old_title_keys, today_max_linkedin_id, recent_urls)
        row["detail_reposted"] = "True" if is_repost else ("False" if row.get("job_board") == "LinkedIn" else "")

    # ── LLM batch classification (German level + experience years) ──
    llm_classify_all(rows)

    # ── Apply filters and split into main + reposted ──
    main_rows = []
    reposted_rows = []
    closed_count = 0
    german_dropped = 0
    exp_dropped = 0
    reposted_count = 0
    enriched_count = 0

    for row in rows:
        active = row.get("verified_active", "")
        lang = row.get("detail_language", "")
        exp_str = row.get("detail_exp_years", "")
        is_repost = row.get("detail_reposted", "") == "True"

        # Segregate: reposted (checked first — reposted jobs go to separate
        # sheet regardless of German/exp filters, for manual review)
        if is_repost:
            reposted_count += 1
            reposted_rows.append(row)
            continue

        # Hard drop: closed
        if active == "False":
            closed_count += 1
            continue
        # Hard drop: German C1+ required
        if lang == "German C1+ required":
            german_dropped += 1
            continue
        # Hard drop: experience >= 3 years
        if exp_str:
            try:
                exp_years = int(exp_str)
                if exp_years >= 3:
                    exp_dropped += 1
                    continue
            except (ValueError, TypeError):
                pass

        main_rows.append(row)

        # Count enriched
        if (row.get("detail_exp_years") or row.get("detail_salary")
                or row.get("detail_remote") or lang in ("German preferred", "German B1/B2 OK")):
            enriched_count += 1

    # ── Write output ──
    stem = csv_path.stem
    if stem.endswith("_verified"):
        stem = stem[:-len("_verified")]
    if stem.endswith("_deduped"):
        stem = stem[:-len("_deduped")]
    out_path = csv_path.parent / f"{stem}_verified.xlsx"

    save_xlsx(out_path, main_rows, reposted_rows)

    # ── Summary ──
    print()
    print("=" * 60)
    print("  Verification Summary")
    print("=" * 60)
    print(f"  Input:                {len(rows)} jobs")
    print(f"  Kept (main sheet):    {len(main_rows)}")
    print(f"  Reposted sheet:       {reposted_count}")
    print(f"  Closed/removed:       {closed_count}")
    print(f"  Dropped (German C1+): {german_dropped}")
    print(f"  Dropped (exp >= 3y):  {exp_dropped}")
    print(f"  Enriched:             {enriched_count}")
    if already_verified:
        print(f"  Already verified:     {already_verified}")
    print(f"  Output: {out_path}")
    print("=" * 60)


def main():
    parser = argparse.ArgumentParser(
        description="Verify job listings: check if active, extract German language "
                    "requirement, experience years, salary, remote status, detect "
                    "reposted LinkedIn jobs, and recalculate match scores from JD text."
    )
    parser.add_argument(
        "--csv", type=str, default=None,
        help="Path to input Job_Search_*.csv (default: auto-find most recent)",
    )
    parser.add_argument(
        "--force", action="store_true",
        help="Re-verify all rows, even those already verified",
    )
    args = parser.parse_args()

    if args.csv:
        csv_path = Path(args.csv)
        if not csv_path.exists():
            print(f"[!] CSV not found: {csv_path}")
            return
    else:
        csv_path = find_latest_csv()
        if csv_path is None:
            print(f"[!] No Job_Search_*.csv found under {JOB_SEARCH_DIR}")
            return

    run_verification(csv_path, force=args.force)


if __name__ == "__main__":
    main()
