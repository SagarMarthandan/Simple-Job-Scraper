#!/usr/bin/env python3
"""
Standalone cleanup tool: removes cross-run duplicates from EXISTING job search sheets.

The cross-run dedup integrated into apify_job_search.py only affects FUTURE runs.
This tool retroactively cleans past sheets by removing jobs that already appeared
in an earlier run's sheet.

Usage:
    python3 dedup_existing_sheets.py              # clean all sheets (oldest kept, newest deduped)
    python3 dedup_existing_sheets.py --dry-run    # report only, no changes
    python3 dedup_existing_sheets.py --from 2026-08-15  # only clean sheets from this date onward

Strategy (same as the integrated dedup):
  - URL key: normalize_job_url() across ALL older runs — exact identity, never expires
  - Title key: normalize_key(company, title) from the SINGLE most recent older run only
    — catches LinkedIn re-lists (new URL per run) and cross-platform dups

For each dated folder (processed oldest→newest), jobs already seen by any OLDER run
are removed from the CSV, JSON, XLSX, and MD files.
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from datetime import datetime

JOB_SEARCH_DIR = Path("/home/sagar/Skills/Jobscraper/Job Search")

FIELDNAMES = ["language", "job_board", "role_type", "title", "company", "location",
              "posted_at", "exp_required", "match_score", "job_url"]


def normalize_key(company: str, title: str) -> str:
    """Generate deduplication key (same as pipeline)."""
    company = company or ""
    title = title or ""
    clean_company = re.sub(r"\b(gmbh|ag|inc|ltd|co|kg|se|corp|llc)\b", "", company, flags=re.IGNORECASE)
    clean_company = re.sub(r"[^\w\s]", "", clean_company).strip().lower()
    clean_title = re.sub(r"[^\w\s]", "", title).strip().lower()
    return f"{clean_company}::{clean_title}"


def normalize_job_url(url: str) -> str:
    """Normalize job URL (same as pipeline). LinkedIn: drop tracking params."""
    url = (url or "").strip()
    if not url:
        return ""
    if "linkedin.com" in url:
        url = url.split("?", 1)[0]
    return url.rstrip("/").lower()


def load_run_csv(csv_path: Path) -> list[dict]:
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def rewrite_csv(csv_path: Path, jobs: list[dict]) -> None:
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        for j in jobs:
            writer.writerow({k: j.get(k, "") for k in FIELDNAMES})


def rewrite_json(json_path: Path, jobs: list[dict]) -> None:
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(jobs, f, indent=2, ensure_ascii=False)


def rewrite_md(md_path: Path, jobs: list[dict], now_str: str) -> None:
    report_md = f"""# Daily Job Openings Report (Last 24 Hours)

**Generated:** {now_str} (retroactively deduped)
**CSV File:** `{md_path.parent / md_path.name.replace('.md', '.csv')}` (Sortable in Excel / Calc)

## Constraints Summary
- **Freshness:** Last 24 Hours
- **Experience:** <= 2 Years (No Senior/Lead/Manager roles)
- **Role Types:** Full-Time & Internships (Germany-wide) | Working Student (**Hamburg & Kiel ONLY**)
- **Target Roles:** Data Engineer, Analytics Engineer, Data Analyst, AI Data Engineer

| # | Language | Job Board | Role Type | Company | Job Title | Location | Match | Direct Link |
|---|---|---|---|---|---|---|---|---|
"""
    for idx, j in enumerate(jobs, 1):
        report_md += (f"| {idx} | {j.get('language','')} | **{j.get('job_board','')}** | "
                      f"{j.get('role_type','')} | {j.get('company','')} | {j.get('title','')} | "
                      f"{j.get('location','')} | {j.get('match_score','')} | "
                      f"[Apply Link]({j.get('job_url','')}) |\n")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(report_md)


def rewrite_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    """Regenerate XLSX from the cleaned CSV."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  [!] openpyxl not installed — skipping XLSX regeneration")
        return

    rows = load_run_csv(csv_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Search Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")

    for col, name in enumerate(FIELDNAMES, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, j in enumerate(rows, 2):
        for col, name in enumerate(FIELDNAMES, 1):
            val = j.get(name, "")
            if name == "match_score":
                try:
                    val = int(re.search(r"(\d+)", val).group(1)) if val else 0
                except (AttributeError, ValueError):
                    val = 0
            cell = ws.cell(row=row_idx, column=col, value=val)
            if name == "job_url" and val:
                cell.hyperlink = val
                cell.font = link_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, len(FIELDNAMES) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or ""))
                      for r in range(1, ws.max_row + 1))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 2, 60)

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="Remove cross-run duplicates from existing job search sheets.")
    parser.add_argument("--dry-run", action="store_true", help="Report only, don't modify files")
    parser.add_argument("--from", dest="from_date", metavar="YYYY-MM-DD",
                        help="Only clean sheets from this date onward")
    args = parser.parse_args()

    # Collect dated run folders sorted oldest→newest
    run_folders = []
    for folder in sorted(JOB_SEARCH_DIR.iterdir()):
        if not folder.is_dir():
            continue
        try:
            datetime.strptime(folder.name, "%Y-%m-%d")
        except ValueError:
            continue
        if args.from_date and folder.name < args.from_date:
            continue
        run_folders.append(folder)

    if not run_folders:
        print("No dated run folders found.")
        return

    print(f"Found {len(run_folders)} run folders ({run_folders[0].name} → {run_folders[-1].name})")
    if args.dry_run:
        print("[DRY RUN — no files will be modified]\n")

    # Accumulate keys from older runs as we process newest
    url_keys = set()
    total_removed = 0
    total_kept = 0

    for i, folder in enumerate(run_folders):
        csv_files = sorted(folder.glob("Job_Search_*.csv"))
        if not csv_files:
            continue

        # Title keys from the single most recent OLDER run (i-1)
        title_keys = set()
        if i > 0:
            older_csvs = sorted(run_folders[i - 1].glob("Job_Search_*.csv"))
            if older_csvs:
                for r in load_run_csv(older_csvs[0]):
                    title_keys.add(normalize_key(r.get("company", ""), r.get("title", "")))

        for csv_path in csv_files:
            jobs = load_run_csv(csv_path)
            original_count = len(jobs)

            kept = []
            removed = 0
            for j in jobs:
                url = normalize_job_url(j.get("job_url", ""))
                tk = normalize_key(j.get("company", ""), j.get("title", ""))
                if url and url in url_keys:
                    removed += 1
                    continue
                if tk in title_keys:
                    removed += 1
                    continue
                kept.append(j)
                if url:
                    url_keys.add(url)

            total_removed += removed
            total_kept += len(kept)

            status = "DRY RUN" if args.dry_run else "CLEANED"
            print(f"  {folder.name}: {original_count} → {len(kept)} ({removed} removed) [{status}]")

            if removed > 0 and not args.dry_run:
                rewrite_csv(csv_path, kept)

                # Regenerate JSON
                json_path = csv_path.with_suffix(".json")
                if json_path.exists():
                    rewrite_json(json_path, kept)

                # Regenerate MD
                md_path = folder / "JOB_OPENINGS_LAST_24H.md"
                if md_path.exists():
                    rewrite_md(md_path, kept, f"{folder.name} (retroactively deduped)")

                # Regenerate XLSX
                xlsx_path = csv_path.with_suffix(".xlsx")
                if xlsx_path.exists():
                    rewrite_xlsx(csv_path, xlsx_path)
            elif removed == 0:
                # Still add URLs to the accumulator even if no removals
                pass

    print(f"\nTotal: {total_removed} duplicates removed, {total_kept} jobs kept across {len(run_folders)} runs")
    if args.dry_run:
        print("\nRe-run without --dry-run to apply changes.")


if __name__ == "__main__":
    main()
